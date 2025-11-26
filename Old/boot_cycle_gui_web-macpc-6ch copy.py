"""
Boot Cycle Logger - Enhanced Edition
6-Channel Video Feed Monitoring and State Detection System

Features:
- Real-time monitoring of 6 video channels (3 consoles, 2 outputs each)
- State detection: Scope Connected, Scope Disconnected, No Signal, Other
- Equipment metadata logging (Console serials, Scope IDs)
- Enhanced CSV reporting with cycle timing analysis
- Smart OBS Virtual Camera auto-detection
- Cross-platform camera connection controls
- Comprehensive test reports with statistical analysis

How to run:
  macOS:   source <venv>/bin/activate && python boot_cycle_gui_web-macpc-6ch.py
  Windows: <venv>\\Scripts\\Activate && python boot_cycle_gui_web-macpc-6ch.py
  Windows: Double-click run_windows.bat (auto-setup)
  Windows: Right-click run_windows.ps1 → Run with PowerShell
Then open http://localhost:5055/

Build (Windows):
  .\build-win.ps1
  OR
  pyinstaller --noconfirm --onefile --noconsole ^
    --name BootCycleLogger ^
    --add-data "art;art" ^
    --add-data "templates;templates" ^
    boot_cycle_gui_web-macpc-6ch.py

Build (macOS):
  ./build-mac.sh
  OR
  pyinstaller --noconfirm --onefile --noconsole \
    --name BootCycleLogger-mac \
    --add-data "art:art" \
    --add-data "templates:templates" \
    boot_cycle_gui_web-macpc-6ch.py
"""

import os, sys, threading, time, csv, platform, webbrowser, subprocess

# Excel export support (optional dependency)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[WARNING] openpyxl not installed. Excel export disabled. Install with: pip install openpyxl")

# ---- Ensure chosen port is free (best-effort cross‑platform) ----
def _free_port(port:int):
    """
    Try to kill any process that is currently LISTENing on the given TCP port.
    Best-effort, safe to call even if nothing is bound. Intended to clean up
    previously orphaned Boot Cycle Logger instances.
    """
    try:
        if platform.system() in ("Darwin", "Linux"):
            # Use lsof to find PIDs bound to the port, then kill them.
            p = subprocess.run(["lsof", "-ti", f"tcp:{port}"], capture_output=True, text=True)
            pids = [pid.strip() for pid in (p.stdout or "").splitlines() if pid.strip()]
            for pid in pids:
                try:
                    subprocess.run(["kill", "-9", pid], check=False)
                except Exception:
                    pass
        elif platform.system() == "Windows":
            # Use PowerShell to get the owning process of the local listening port and kill it.
            ps = f"(Get-NetTCPConnection -LocalPort {port} -State Listen).OwningProcess"
            p = subprocess.run(["powershell", "-NoProfile", "-Command", ps], capture_output=True, text=True)
            pids = [pid.strip() for pid in (p.stdout or "").split() if pid.strip().isdigit()]
            for pid in pids:
                try:
                    subprocess.run(["taskkill", "/PID", pid, "/F"], check=False)
                except Exception:
                    pass
    except Exception:
        # As a fallback we do nothing; app.run will error if still occupied.
        pass

# ---- Port selection helpers and global ----
import socket

def _is_port_free(port:int) -> bool:
    """Return True if TCP port is free on localhost."""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(0.2)
        try:
            return s.connect_ex(("127.0.0.1", port)) != 0
        except Exception:
            return False

def _choose_port(base:int=5055, upper:int=5070) -> int:
    """
    Try to free and return a port starting at `base`. If still occupied,
    scan upward to `upper` and pick the first free one.
    """
    # First, try to free the base port and reuse it
    _free_port(base)
    if _is_port_free(base):
        return base
    # Otherwise find the next free port
    for p in range(base + 1, upper + 1):
        if _is_port_free(p):
            return p
    # Fallback: return base even if busy (will raise at app.run)
    return base

# Chosen HTTP port (set in __main__)
PORT = 5055
from datetime import datetime
from flask import Flask, jsonify, request, render_template_string, render_template, Response, send_file
import cv2, numpy as np
from PIL import Image
import imagehash as ih
import subprocess

# Video source & capture options (sane defaults, then OS tweaks)
BACKEND    = "auto"     # auto|MSMF|DSHOW|AVFOUNDATION|V4L2
FOURCC     = "auto"     # auto|MJPG|YUY2
RES_PRESET = "1080p"    # 1080p|720p
SRC        = "0"        # default camera index; macOS override below

if platform.system() == "Darwin":
    # OBS/USB capture commonly shows up as index 1 on macOS
    SRC = "1"

if platform.system() == "Windows":
    # These help MSMF behave with MJPG/YUY2 and some H.264 sources
    os.environ.setdefault("OPENCV_VIDEOIO_MSMF_ENABLE_HW_TRANSFORMS", "0")
    os.environ.setdefault("OPENCV_VIDEOIO_MSMF_ENABLE_H264", "1")

# ---- Resolve app directory for data files (works in PyInstaller and from source) ----
APP_DIR = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))

TEMPLATE_DIR = os.path.join(APP_DIR, "templates")
STATIC_DIR   = os.path.join(APP_DIR, "static")

# Art folder path (ROI connected reference)
ART_DIR = os.path.join(APP_DIR, "art")

# Logs folder root (works from source and PyInstaller)
LOG_ROOT = os.path.join(APP_DIR, "logs")

# Helper to stringify FOURCC integer to string
def _fourcc_to_str(v:int) -> str:
    try:
        return ''.join([chr((int(v) >> (8*i)) & 0xFF) for i in range(4)])
    except Exception:
        return ''

# ---------- defaults ----------
CENTER_W = 1920
THRESH   = 10
STABLE   = 1
DARK_MEAN= 22.0
DARK_STD = 12.0

# --- 6‑feed grid layout (3x2) ---
GRID_COLS = 3
GRID_ROWS = 2
GRID_FEEDS = GRID_COLS * GRID_ROWS  # 6

# Reference images + CSV
BARS_REF = os.path.join(ART_DIR, "Scope-Disconnected.png")  # keep your disconnected image (now relative to art/)
# ROI connected reference inside /art (used for INTERFACE detection)
INT_REF  = os.path.join(ART_DIR, "Boot-Reliabilty-Testing.png")
CSV_PATH = os.path.join(LOG_ROOT, f"boot_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")

# Additional INT reference discovery (multiple "Boot-Reliabilty-Testing*" variants)
INT_REF_PREFIX = "Boot-Reliabilty-Testing"   # (spelling matches your files)
INT_REF_EXTS   = (".png", ".jpg", ".jpeg", ".bmp", ".webp")



# Anti-flicker defaults (Windows virtual cam can be jittery)
HOLD_MS  = 800   # minimum time a new state must persist before we accept it
MARGIN   = 2     # hysteresis margin for phash distance thresholds

# Throttle how often we update the live crop to avoid pushing detector around
THUMB_EVERY_MS = 1500

# Live thumbnail target size (keep 16:9 to match source aspect)
THUMB_W = 256
THUMB_H = 144

def _placeholder_thumb(width: int = 780, height: int = 288):
    """Generate a neutral placeholder when no frame is available."""
    img = np.zeros((height, width, 3), dtype=np.uint8)
    # subtle border
    cv2.rectangle(img, (0, 0), (width - 1, height - 1), (60, 60, 60), 1)
    try:
        cv2.putText(img, "no frame", (12, height // 2), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (180, 180, 180), 1, cv2.LINE_AA)
    except Exception:
        pass
    return img

def crop(bgr, cw=CENTER_W):
    h, w, _ = bgr.shape
    # normalize/validate cw
    try:
        if cw is None:
            cw = CENTER_W
        cw = int(cw)
    except Exception:
        cw = CENTER_W

    # If cw is <= 0 or >= frame width, use the full frame (no side-gutter crop).
    if cw <= 0 or cw >= w:
        gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
        return cv2.resize(gray, (THUMB_W, THUMB_H), interpolation=cv2.INTER_AREA)

    # Compute left/right gutter width
    s = (w - cw) // 2
    if s <= 0:
        gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
        return cv2.resize(gray, (THUMB_W, THUMB_H), interpolation=cv2.INTER_AREA)

    # Concatenate only the side gutters, convert to gray, resize to thumbnail
    side = np.concatenate([bgr[:, :s], bgr[:, w - s:]], axis=1)
    gray = cv2.cvtColor(side, cv2.COLOR_BGR2GRAY)
    return cv2.resize(gray, (THUMB_W, THUMB_H), interpolation=cv2.INTER_AREA)

# ---------- ROI detector for "Device Connected" ----------
# 6‑feed grid ROI: each tile is ~640x540 when the full frame is 1920x1080.
# Target rectangle per tile: x≈140, y≈340 (top-left), w≈133, h≈200.
GRID_BASE_W, GRID_BASE_H = 640, 540
ROI_TILE_BASE = dict(x=140, y=GRID_BASE_H - 200, w=133, h=200)  # 140,340,133,200

ROI_FRAC = dict(
    x = ROI_TILE_BASE["x"] / float(GRID_BASE_W),
    y = ROI_TILE_BASE["y"] / float(GRID_BASE_H),
    w = ROI_TILE_BASE["w"] / float(GRID_BASE_W),
    h = ROI_TILE_BASE["h"] / float(GRID_BASE_H),
)

def _roi_box_for_frame(w:int, h:int, inset:int=0):
    """
    Bottom-anchored ROI for each tile using base pixels:
    x≈140, w≈133 scaled from 640-wide tiles; height≈200 scaled from 540-tall tiles.
    The ROI sits flush on the bottom regardless of rounding differences.
    """
    # Horizontal scale against 640; vertical against 540
    rw = max(1, int(round(ROI_TILE_BASE["w"] * (w / float(GRID_BASE_W)))))
    rh = max(1, int(round(ROI_TILE_BASE["h"] * (h / float(GRID_BASE_H)))))
    x  = int(round(ROI_TILE_BASE["x"] * (w / float(GRID_BASE_W))))
    y  = max(0, h - rh)  # bottom anchor

    # Clamp
    if x + rw > w: x = max(0, w - rw)
    if y + rh > h: y = max(0, h - rh)

    inset = max(0, int(inset))
    if inset:
         rw = max(1, rw - 2*inset)
         rh = max(1, rh - 2*inset)
         x  = min(max(0, x + inset), max(0, w - rw))
         y0 = y + inset  # keep bottom roughly anchored
         y  = min(max(0, y0), max(0, h - rh))
    return x, y, rw, rh

def roi_connected_gray(bgr):
    """Extract the INTERFACE ROI region, grayscale it, return a square for stable pHash."""
    h, w = bgr.shape[:2]
    x, y, rw, rh = _roi_box_for_frame(w, h, getattr(mon, "roi_inset_px", 0))
    roi = bgr[y:y+rh, x:x+rw]
    if roi.size == 0:
        roi = bgr
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    return cv2.resize(gray, (256, 256), interpolation=cv2.INTER_AREA)

# --- Grid helpers ---
def split_grid(frame, cols=GRID_COLS, rows=GRID_ROWS):
    """Return (tiles, rects). rects are (x,y,w,h) in the original frame."""
    h, w = frame.shape[:2]
    tile_w = w // cols
    tile_h = h // rows
    tiles, rects = [], []
    for r in range(rows):
        for c in range(cols):
            x = c * tile_w
            y = r * tile_h
            w_c = (w - x) if c == cols - 1 else tile_w
            h_r = (h - y) if r == rows - 1 else tile_h
            rects.append((x, y, w_c, h_r))
            tiles.append(frame[y:y+h_r, x:x+w_c])
    return tiles, rects

def draw_status_badge(img, x, y, text, color):
    """Draw a simple status badge rectangle with text."""
    pad_x, pad_y = 10, 8
    font = cv2.FONT_HERSHEY_SIMPLEX
    scale = 0.65
    thick = 2
    (tw, th), _ = cv2.getTextSize(text, font, scale, thick)
    w = tw + pad_x*2
    h = th + pad_y*2
    cv2.rectangle(img, (x, y), (x+w, y+h), color, -1)
    cv2.rectangle(img, (x, y), (x+w, y+h), (0,0,0), 1)
    cv2.putText(img, text, (x+pad_x, y+pad_y+th-4), font, scale, (255,255,255), thick, cv2.LINE_AA)
    return (x+w, y+h)

# --- ROI brightness statistics helper ---
def roi_stats(gray):
    """
    Return (mean_luma, std_luma, bright_frac) for an 8‑bit grayscale ROI.
    bright_frac is the fraction of pixels > 240 (very bright/white).
    """
    m = float(np.mean(gray))
    s = float(np.std(gray))
    # fraction of nearly-white pixels
    bright_frac = float((gray > 240).mean())
    return m, s, bright_frac

# --- Compute average ROI stats from reference images ---
def ref_roi_stats_from_paths(paths):
    """
    Compute average ROI brightness stats across all connected reference images.
    Returns (ref_mean, ref_std, ref_bright) or (None, None, None) if none available.
    """
    means, stds, brights = [], [], []
    for p in (paths or []):
        try:
            img = cv2.imread(p)
            if img is None:
                continue
            roi = roi_connected_gray(img)
            m, s, b = roi_stats(roi)
            means.append(m); stds.append(s); brights.append(b)
        except Exception:
            continue
    if not means:
        return None, None, None
    return float(np.mean(means)), float(np.mean(stds)), float(np.mean(brights))

def _effective_cw_for_width(width:int, user_cw:int) -> int:
    """
    Use side-gutter crop for bars even when the UI center width is >= frame width.
    If user_cw is invalid or >= width, fall back to ~60% of the frame width.
    """
    try:
        c = int(user_cw)
    except Exception:
        c = CENTER_W
    if c <= 0 or c >= int(width):
        return max(1, int(round(width * 0.60)))
    return c

def _equalize_hist(gray):
    """Apply histogram equalization to a grayscale image."""
    return cv2.equalizeHist(gray)

def ph_int_ref(path):
    """pHash of the ROI (white 400x400 square area) from a single INT reference image.
    Crops to ROI and applies histogram equalization before hashing.
    """
    img = cv2.imread(path)
    if img is None:
        raise RuntimeError(f"Failed to read INT reference image: {path}")
    roi_gray = roi_connected_gray(img)
    roi_eq = _equalize_hist(roi_gray)
    return ih.phash(Image.fromarray(roi_eq))

def ph_int_ref_list(paths):
    """pHashes of the ROI from all provided INT reference images (skip unreadable).
    Crops to ROI and applies histogram equalization before hashing.
    """
    hashes = []
    for p in paths or []:
        try:
            img = cv2.imread(p)
            if img is None:
                continue
            roi_gray = roi_connected_gray(img)
            roi_eq = _equalize_hist(roi_gray)
            hashes.append(ih.phash(Image.fromarray(roi_eq)))
        except Exception:
            continue
    return hashes

   

def ph(gray_crop):
    return ih.phash(Image.fromarray(gray_crop))

def ph_ref(path, cw):
    img = cv2.imread(path)
    if img is None:
        raise RuntimeError(f"Failed to read reference image: {path} (cwd={os.getcwd()})")
    ref_w = img.shape[1]
    cw_eff = _effective_cw_for_width(ref_w, cw)
    return ph(crop(img, cw_eff))

# --- Helper: dual bars hashes (side-gutters crop and full-frame) ---

# --- Helper: pHash of the ROI region from the BARS (Scope-Disconnected) reference image ---
def ph_bars_ref_roi(path):
    """pHash of the ROI region taken from the BARS (disconnected) reference image."""
    img = cv2.imread(path)
    if img is None:
        raise RuntimeError(f"Failed to read BARS reference image: {path}")
    roi_gray = roi_connected_gray(img)
    roi_eq = _equalize_hist(roi_gray)
    return ih.phash(Image.fromarray(roi_eq))

def discover_int_refs():
    """
    Return a de-duplicated, ordered list of INT reference paths. It prefers the
    canonical INT_REF first and then any files in art/ that start with
    "Boot-Reliabilty-Testing" (any common image extension).
    """
    paths = []
    try:
        if os.path.exists(INT_REF):
            paths.append(INT_REF)
        for name in os.listdir(ART_DIR):
            low = name.lower()
            if low.startswith(INT_REF_PREFIX.lower()) and low.endswith(INT_REF_EXTS):
                p = os.path.join(ART_DIR, name)
                if os.path.exists(p):
                    paths.append(p)
    except Exception:
        pass
    # de-dup while preserving order
    out, seen = [], set()
    for p in paths:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out

def label_for(state):
    """Map internal states to user-facing labels (used in UI and CSV)."""
    if state == "BARS":
        return "Scope Disconnected"
    elif state == "INTERFACE":
        return "Scope Connected"
    elif state == "NO_SIGNAL":
        return "No Signal"
    else:
        return "Other"

class Monitor:
    def __init__(self):
        # backend selection (string name; resolved at open time)
        self.backend_name = BACKEND  # "auto", "MSMF", "DSHOW", "AVFOUNDATION", "V4L2"
        self.fourcc = FOURCC
        self.res_preset = RES_PRESET

        # config
        self.src       = SRC
        self.stream_path = ""               # optional URL; overrides src if set
        self.center_w  = CENTER_W
        self.thresh    = THRESH
        self.stable_frames = STABLE
        self.dark_mean = DARK_MEAN
        self.dark_std  = DARK_STD
        # ROI brightness override controls (None => adaptive)
        self.white_frac_gate = 1  # fraction 0.0..1.0 or % from UI (default: 1 = 1%)
        self.mean_gate       = 1  # grayscale mean threshold (0..255) (default: 1)
        self.hold_ms   = HOLD_MS
        self.margin    = MARGIN
        self.roi_inset_px = 1  # shrink ROI inward on all sides (px in tile space) (default: 1)
        # live-crop thumbnail throttling
        self.thumb_every_ms = THUMB_EVERY_MS
        self.thumb_enabled  = True
        self._last_thumb_ts = 0.0
        self.bars_ref  = BARS_REF
        self.int_ref   = INT_REF
        self.int_ref_paths = discover_int_refs()
        # int_ref2 removed/not used
        self.csv_path  = CSV_PATH

        # Equipment metadata: 3 consoles, each mapped to 2 video channels
        # Console 1: Videos 1&2, Console 2: Videos 3&4, Console 3: Videos 5&6
        self.equipment = {
            "console1": {"serial": "", "scope_id": "", "videos": [1, 2]},
            "console2": {"serial": "", "scope_id": "", "videos": [3, 4]},
            "console3": {"serial": "", "scope_id": "", "videos": [5, 6]}
        }
        self.test_start_time = None  # Timestamp when test actually starts

        # runtime state
        self.lock = threading.Lock()
        self.video_running = False  # Video feed is running (display only)
        self.detection_active = False  # Detection/logging is active
        self.running = False  # Legacy flag for compatibility
        self.status = "idle"
        self.worker = None  # thread running the capture loop
        self.last_cfg = None
        self.count_bars = 0
        self.count_int  = 0
        self.count_other= 0
        self.cycles     = 0
        # per-tile (6-ch) state/counters
        self.await_connect = True
        self.tile_last   = ["UNKNOWN"] * GRID_FEEDS
        self.tile_counts = [0] * GRID_FEEDS  # increments when a tile transitions to INTERFACE
        self.tile_disconnected_counts = [0] * GRID_FEEDS  # increments when transitioning FROM BARS to NO_SIGNAL (complete disconnect cycle)
        
        # Enhanced per-channel tracking for timing and cycle analysis
        self.tile_disconnected_start = [None] * GRID_FEEDS  # timestamp when Scope Disconnected started
        self.tile_cycle_times = [[] for _ in range(GRID_FEEDS)]  # list of elapsed times for each channel
        self.tile_complete_cycles = [0] * GRID_FEEDS  # count of complete cycles (increments on each INTERFACE)
        self.tile_partial_cycles = [0] * GRID_FEEDS  # count of partial cycles: NO_SIGNAL → BARS → NO_SIGNAL
        self.tile_state_history = [[] for _ in range(GRID_FEEDS)]  # track state progression for each channel
        self.tile_in_partial_cycle = [False] * GRID_FEEDS  # track if we're in a potential partial cycle (entered BARS from NO_SIGNAL)

        # stabilization
        self._last = "UNKNOWN"
        self._raw_last = "UNKNOWN"
        self._raw_stable = 0
        self._last_change_ts = 0.0

        # last-seen timestamps
        self.last_seen = {"BARS": None, "INTERFACE": None, "OTHER": None, "NO_SIGNAL": None}

        # live frames/metrics for thumbnail + tester
        self.last_frame = None
        self.last_crop  = None
        self.last_metrics = {"db": None, "di": None, "mean": None, "std": None}
        self.last_grid = None  # cached per-frame grid detections for UI consistency

        # cached refs for quick /thumb grid annotation
        self._bars_h_roi = None
        self._int_h_list = None
        self._ref_mean = None
        self._ref_bright = None

    def _resolve_backend(self):
        name = (self.backend_name or "auto").strip().lower()
        if name == "auto":
            if platform.system() == "Windows":
                return cv2.CAP_MSMF
            elif platform.system() == "Darwin":
                return cv2.CAP_AVFOUNDATION
            else:
                return cv2.CAP_V4L2
        if name == "msmf": return cv2.CAP_MSMF
        if name == "dshow": return cv2.CAP_DSHOW
        if name == "avfoundation": return cv2.CAP_AVFOUNDATION
        if name == "v4l2": return cv2.CAP_V4L2
        # fallback
        if platform.system() == "Windows": return cv2.CAP_MSMF
        if platform.system() == "Darwin": return cv2.CAP_AVFOUNDATION
        return cv2.CAP_V4L2

    def reset_counts_and_roll_csv(self):
        self.count_bars = self.count_int = self.count_other = self.cycles = 0
        self._last = "UNKNOWN"
        self._raw_last = "UNKNOWN"
        self._raw_stable = 0
        self.last_seen = {"BARS": None, "INTERFACE": None, "OTHER": None, "NO_SIGNAL": None}
        # new timestamped CSV on each clear
        self.csv_path = os.path.join(LOG_ROOT, f"boot_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        self.await_connect = True
        # Reset per-tile counters
        self.tile_last = ["UNKNOWN"] * GRID_FEEDS
        self.tile_counts = [0] * GRID_FEEDS
        self.tile_disconnected_counts = [0] * GRID_FEEDS
        self.tile_disconnected_start = [None] * GRID_FEEDS
        self.tile_cycle_times = [[] for _ in range(GRID_FEEDS)]
        self.tile_complete_cycles = [0] * GRID_FEEDS  # Reset complete cycle counts
        self.tile_partial_cycles = [0] * GRID_FEEDS  # Reset partial cycle counts
        self.tile_in_partial_cycle = [False] * GRID_FEEDS
        self.tile_state_history = [[] for _ in range(GRID_FEEDS)]

    def reset_tallies(self):
        """Zero counters and last-seen timestamps without rolling the CSV file."""
        self.count_bars = 0
        self.count_int = 0
        self.count_other = 0
        self.cycles = 0
        self._last = "UNKNOWN"
        self._raw_last = "UNKNOWN"
        self._raw_stable = 0
        self.last_seen = {"BARS": None, "INTERFACE": None, "OTHER": None, "NO_SIGNAL": None}
        self.await_connect = True
        self.tile_last   = ["UNKNOWN"] * GRID_FEEDS
        self.tile_counts = [0] * GRID_FEEDS
        self.tile_disconnected_counts = [0] * GRID_FEEDS
        self.tile_disconnected_start = [None] * GRID_FEEDS
        self.tile_cycle_times = [[] for _ in range(GRID_FEEDS)]
        self.tile_complete_cycles = [0] * GRID_FEEDS
        self.tile_partial_cycles = [0] * GRID_FEEDS
        self.tile_in_partial_cycle = [False] * GRID_FEEDS
        self.tile_state_history = [[] for _ in range(GRID_FEEDS)]

mon = Monitor()

def _get_excel_path(csv_path):
    """Get Excel file path from CSV path."""
    return csv_path.replace('.csv', '.xlsx')

def _create_excel_file(excel_path):
    """Create Excel file with formatted headers (frozen top row, bold)."""
    if not EXCEL_AVAILABLE:
        return None
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Boot Cycle Log"
        
        # Column headers
        headers = [
            "Timestamp",
            "Video Channel",
            "Console Serial",
            "Scope ID",
            "State",
            "Elapsed Secs",
            "full cycle count",
            "partial cycle count",
            "Disconnected Match Distance",
            "Connected Match Distance",
            "Event Type"
        ]
        
        # Write headers with bold formatting
        bold_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font
        
        # Freeze top row
        ws.freeze_panes = "A2"
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
        
        wb.save(excel_path)
        print(f"[Excel] Created new file with formatted headers: {excel_path}")
        return excel_path
    except Exception as e:
        print(f"[Excel] Error creating file: {e}")
        return None

def _append_to_excel(excel_path, row_data):
    """Append a row to the Excel file."""
    if not EXCEL_AVAILABLE or not excel_path:
        return
    
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        ws.append(row_data)
        wb.save(excel_path)
    except Exception as e:
        print(f"[Excel] Error appending row: {e}")

def open_csv(path):
    """Open CSV file and write column headers if file is new/empty. Also creates Excel file."""
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    
    # Check if file exists and is empty
    file_is_new = not os.path.exists(path) or os.path.getsize(path) == 0
    
    f = open(path, "a", newline="")
    w = csv.writer(f)
    
    if file_is_new:
        # Write column headers for new CSV file
        headers = [
            "Timestamp",
            "Video Channel",
            "Console Serial",
            "Scope ID",
            "State",
            "Elapsed Secs",
            "full cycle count",
            "partial cycle count",
            "Disconnected Match Distance",
            "Connected Match Distance",
            "Event Type"
        ]
        w.writerow(headers)
        f.flush()  # Ensure headers are written immediately
        print(f"[CSV] Created new file with headers: {path}")
        
        # Also create Excel file with formatted headers
        if EXCEL_AVAILABLE:
            excel_path = _get_excel_path(path)
            _create_excel_file(excel_path)
    
    return f, w

def _get_equipment_for_video(video_num):
    """Get console serial and scope ID for a given video channel (1-6)."""
    with mon.lock:
        for console_key, info in mon.equipment.items():
            if video_num in info["videos"]:
                return info["serial"], info["scope_id"]
    return "", ""

def _csv_append(video:int, state:str, elapsed_secs=None, cycle_num=None, partial_cycle_num=None, bars_dist=None, int_dist=None, event_type:str="state_change"):
    """
    Append a single event row to the current CSV and Excel file with enhanced format.
    Only logs when detection is active.
    Columns: Timestamp, Video Channel, Console Serial, Scope ID, State, Elapsed Secs, 
             full cycle count, partial cycle count, Disconnected Match Distance, Connected Match Distance, Event Type
    
    Match distances are perceptual hash (pHash) distances comparing current frame to reference images.
    Lower values = closer match. Typical range: 0-20 (0=exact match).
    
    Note: elapsed_secs is not logged for NO_SIGNAL state.
    """
    # Only log when detection is active
    with mon.lock:
        if not mon.detection_active:
            return
    
    try:
        os.makedirs(os.path.dirname(mon.csv_path) or ".", exist_ok=True)
        ts = datetime.now().isoformat(timespec="milliseconds")
        console_serial, scope_id = _get_equipment_for_video(video)
        state_label = label_for(state)
        
        # Don't log elapsed_secs for NO_SIGNAL state
        elapsed_value = None if state == "NO_SIGNAL" else elapsed_secs
        
        # Prepare row data for CSV (strings)
        csv_row_data = [
            ts,
            video,
            console_serial,
            scope_id,
            state_label,
            f"{elapsed_value:.2f}" if elapsed_value is not None else "",
            cycle_num if cycle_num is not None else "",
            partial_cycle_num if partial_cycle_num is not None else "",
            bars_dist if bars_dist is not None else "",
            int_dist if int_dist is not None else "",
            event_type
        ]
        
        # Prepare row data for Excel (numbers as numbers, not strings)
        excel_row_data = [
            ts,
            int(video) if isinstance(video, (int, float)) else video,
            console_serial,
            scope_id,
            state_label,
            float(elapsed_value) if elapsed_value is not None else None,
            int(cycle_num) if cycle_num is not None and cycle_num != "" else None,
            int(partial_cycle_num) if partial_cycle_num is not None and partial_cycle_num != "" else None,
            float(bars_dist) if bars_dist is not None and bars_dist != "" else None,
            float(int_dist) if int_dist is not None and int_dist != "" else None,
            event_type
        ]
        
        # Write to CSV
        with open(mon.csv_path, "a", newline="") as _f:
            _w = csv.writer(_f)
            _w.writerow(csv_row_data)
        
        # Also write to Excel if available
        if EXCEL_AVAILABLE:
            excel_path = _get_excel_path(mon.csv_path)
            if os.path.exists(excel_path):
                _append_to_excel(excel_path, excel_row_data)
            else:
                # Excel file doesn't exist yet, create it
                _create_excel_file(excel_path)
                _append_to_excel(excel_path, excel_row_data)
                
    except Exception as e:
        print(f"[CSV] Error appending: {e}")

def _composite_thumb_from_frame(bgr):
    """
    Composite debug view: ROI (384x384, left) + full frame (native aspect, 384px tall, right) with labels.
    ROI is a 1:1 square, 384x384. Full-frame is resized to height 384, width = int(384 * (w/h)).
    There is a 12px gap between panels.
    """
    ROI_SIZE = 384  # ROI panel size (square)
    GAP = 12
    # Get original frame size
    h, w = bgr.shape[:2]
    # Compute full-frame preview width to maintain aspect ratio at height=384
    full_h = ROI_SIZE
    full_w = int(round(full_h * (w / h)))
    # Output canvas size
    OUT_W = ROI_SIZE + GAP + full_w
    OUT_H = ROI_SIZE

    # Build ROI (384x384)
    try:
        roi_gray = roi_connected_gray(bgr)
        roi_rgb = cv2.cvtColor(roi_gray, cv2.COLOR_GRAY2BGR)
        roi_rgb = cv2.resize(roi_rgb, (ROI_SIZE, ROI_SIZE), interpolation=cv2.INTER_AREA)
    except Exception:
        roi_rgb = np.zeros((ROI_SIZE, ROI_SIZE, 3), dtype=np.uint8)

    left_panel = roi_rgb

    # Build full-frame (native aspect, 384px tall)
    try:
        right_panel = cv2.resize(bgr, (full_w, full_h), interpolation=cv2.INTER_AREA)
    except Exception:
        right_panel = np.zeros((full_h, full_w, 3), dtype=np.uint8)

    # Compose output canvas
    comp = np.zeros((OUT_H, OUT_W, 3), dtype=np.uint8)
    # Place ROI at (0,0)
    comp[0:ROI_SIZE, 0:ROI_SIZE] = left_panel
    # Place gap (GAP px wide, black, nothing needed)
    # Place full-frame at (ROI_SIZE + GAP, 0)
    comp[0:full_h, ROI_SIZE + GAP:ROI_SIZE + GAP + full_w] = right_panel

    # Draw outer border
    cv2.rectangle(comp, (0, 0), (OUT_W - 1, OUT_H - 1), (60, 60, 60), 1)

    # Labels (ROI: top left, Full: above full-frame panel)
    try:
        cv2.putText(comp, "ROI", (16, 32), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (200, 200, 200), 2, cv2.LINE_AA)
        # Place "Full" label above the full-frame panel, left-aligned to its panel
        full_label_x = ROI_SIZE + GAP + 16
        full_label_y = 32
        cv2.putText(comp, "Full", (full_label_x, full_label_y), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (200, 200, 200), 2, cv2.LINE_AA)
    except Exception:
        pass

    return comp

def _mac_try_indices_for_nonblack(indices, backend, want_w=1920, want_h=1080, fourcc_sel="auto"):
    """
    macOS helper: try several AVFoundation indices and return the first capture
    that both opens and yields a non‑black frame (mean luminance > 5).
    Returns (cap, used_index) or (None, None).
    """
    for idx in indices:
        try:
            cap = cv2.VideoCapture(idx, backend)
            if not cap or not cap.isOpened():
                try:
                    if cap: cap.release()
                except Exception:
                    pass
                continue
            try: cap.set(cv2.CAP_PROP_CONVERT_RGB, 1)
            except Exception: pass
            if fourcc_sel == "MJPG":
                try: cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'MJPG'))
                except Exception: pass
            elif fourcc_sel == "YUY2":
                try: cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'YUY2'))
                except Exception: pass
            try:
                cap.set(cv2.CAP_PROP_FRAME_WIDTH,  want_w)
                cap.set(cv2.CAP_PROP_FRAME_HEIGHT, want_h)
            except Exception:
                pass
            # warm up - give camera more time to initialize
            for _ in range(15):
                cap.read()
                time.sleep(0.05)  # 50ms between reads
            ok, frame = cap.read()
            if not ok or frame is None:
                try:
                    cap.release()
                except Exception:
                    pass
                continue
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            mean_val = float(np.mean(gray))
            print(f"[mac_try] index={idx} mean={mean_val:.1f}")
            if mean_val > 5.0:
                print(f"[mac_try] ✓ selected AVFoundation index={idx} (mean={mean_val:.1f})")
                return cap, idx
            # black feed: try next
            print(f"[mac_try] ✗ skipping index={idx} (too dark, mean={mean_val:.1f})")
            try:
                cap.release()
            except Exception:
                pass
        except Exception:
            try:
                if cap: cap.release()
            except Exception:
                pass
            continue
    return None, None

def decide(
    frame,
    bars_h_roi,
    int_h_list,
    center_w,
    thresh,
    dmean,
    dstd,
    margin,
    ref_mean=None,
    ref_bright=None,
    white_frac_override=None,
    mean_gate_override=None,
):
    """
    ROI-only decision with a stricter "white ROI" gate:
      - INTERFACE: (pHash near any Boot-Reliabilty-Testing* ref) AND (ROI is truly white)
                   OR (ROI passes a strong "white ROI" brightness gate)
      - BARS     : ROI pHash near the BARS (Scope-Disconnected) ROI
      - NO_SIGNAL: dark/flat full frame (suppressed if ROI is clearly white)
      - OTHER    : fallback
    Returns:
        (det, db, di, cg, mean_lum, std_lum, roi_mean, roi_bright, bright_ok, white_frac_gray, white_frac_rgb)
        - bright_ok is True only when the ROI is sufficiently white (see gates below).
    """
    # --- Compute ROI (gray and equalized) for pHash ---
    # Smooth slightly to reduce pixel-level noise before stats/pHash.
    roi_plain = roi_connected_gray(frame)
    roi_smooth = cv2.GaussianBlur(roi_plain, (5, 5), 0)
    roi_eq = _equalize_hist(roi_smooth)

    # Also get the color ROI to verify "whiteness" (avoid false positives on colored blocks)
    h, w = frame.shape[:2]
    rx, ry, rw, rh = _roi_box_for_frame(w, h, getattr(mon, "roi_inset_px", 0))
    roi_bgr = frame[ry:ry + rh, rx:rx + rw] if rh > 0 and rw > 0 else frame

    phv_roi_plain = ih.phash(Image.fromarray(roi_plain))
    phv_roi_eq = ih.phash(Image.fromarray(roi_eq))

    # Distance to all INT references (use both plain and equalized to be robust)
    di_list = []
    for hsh in (int_h_list or []):
        try:
            di_list.append(int(phv_roi_eq - hsh))
        except Exception:
            pass
        try:
            di_list.append(int(phv_roi_plain - hsh))
        except Exception:
            pass
    di = min(di_list or [999])

    # Distance to BARS ref (ROI)
    try:
        db = int(phv_roi_eq - bars_h_roi)
    except Exception:
        db = 999

    # --- Luminance stats (on the smoothed ROI) ---
    roi_mean, roi_std, roi_bright = roi_stats(roi_smooth)
    full_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    mean_lum = float(np.mean(full_gray))
    std_lum = float(np.std(full_gray))
    is_dark_flat = (mean_lum < dmean) and (std_lum < dstd)

    # --- Optional: try a slightly tighter ROI if the current mean is a bit low ---
    # This helps when the nominal rectangle grabs a sliver of the black tile border.
    try:
        if roi_mean < 195.0:
            h0, w0 = frame.shape[:2]
            rx, ry, rw, rh = _roi_box_for_frame(w0, h0, getattr(mon, "roi_inset_px", 0) + 8)
            alt = frame[ry:ry + rh, rx:rx + rw] if rh > 0 and rw > 0 else frame
            alt_gray = cv2.cvtColor(alt, cv2.COLOR_BGR2GRAY)
            alt_smooth = cv2.GaussianBlur(alt_gray, (5, 5), 0)
            alt_mean, alt_std, alt_bright = roi_stats(alt_smooth)
            if alt_mean > roi_mean:  # keep the "cleaner" ROI if it improved
                roi_smooth = alt_smooth
                roi_mean, roi_std, roi_bright = alt_mean, alt_std, alt_bright
                roi_bgr = alt
    except Exception:
        pass

    # --- Adaptive "white ROI" gate (tolerant to near‑white) ---
    base_bright_frac = 0.05  # default fraction of near‑white pixels
    base_mean_gate = 65.0    # default minimum ROI mean
    if ref_mean is not None and ref_bright is not None:
        # Soften gates relative to references to allow real-world tolerance.
        base_bright_frac = max(0.02, float(ref_bright) * 0.8)
        base_mean_gate   = max(50.0, float(ref_mean) - 10.0)

    # Optional UI overrides
    if white_frac_override is not None:
        try:
            wf = float(white_frac_override)
            if wf > 1.0:  # treat as percent
                wf = wf / 100.0
            base_bright_frac = max(0.0, min(1.0, wf))
        except Exception:
            pass
    if mean_gate_override is not None:
        try:
            base_mean_gate = float(mean_gate_override)
        except Exception:
            pass

    # Channel thresholds for "near‑white"
    rgb_thresh  = 235
    gray_thresh = 230

    # RGB near‑white fraction: all three channels above threshold.
    try:
        rgb_mask = (
            (roi_bgr[:, :, 0] >= rgb_thresh)
            & (roi_bgr[:, :, 1] >= rgb_thresh)
            & (roi_bgr[:, :, 2] >= rgb_thresh)
        )
        white_frac_rgb = float(np.mean(rgb_mask))
    except Exception:
        white_frac_rgb = 0.0

    # Grayscale near‑white fraction on smoothed ROI
    try:
        white_frac_gray = float(np.mean(roi_smooth >= gray_thresh))
    except Exception:
        white_frac_gray = 0.0

    # Combine signals: accept if either RGB or Gray fraction is sufficient AND mean is high enough.
    frac_gate_rgb  = max(0.015, 0.6 * float(base_bright_frac))
    frac_gate_gray = max(0.03,  0.8 * float(base_bright_frac))
    strong_white = (roi_mean >= base_mean_gate) and (
        (white_frac_rgb  >= frac_gate_rgb) or
        (white_frac_gray >= frac_gate_gray)
    )

    # --- pHash gates (computed above) ---
    bars_gate = thresh + margin
    int_gate  = thresh + margin + 2
    phash_ok  = (di <= int_gate)

    # --- Looser gray-interface allowance ---
    # If the ROI is bright-ish (but not fully white) and the pHash matches,
    # we still consider it connected. This is what lets ~211 means pass.
    loose_gray_gate = max(180.0, float(base_mean_gate) - 30.0)
    gray_ok = (roi_mean >= loose_gray_gate) and phash_ok

    # Final brightness decision used by the UI overlay
    bright_ok = bool(strong_white or gray_ok)

    # --- Decision ---
    if is_dark_flat and not (gray_ok or strong_white):
        det = "NO_SIGNAL"
    elif phash_ok and (strong_white or gray_ok):
        det = "INTERFACE"
    elif strong_white:
        det = "INTERFACE"
    elif db < bars_gate:
        det = "BARS"
    else:
        det = "OTHER"

    # Provide side-gutter crop for legacy metrics/CSV
    cw_eff_runtime = _effective_cw_for_width(frame.shape[1], center_w)
    cg = crop(frame, cw_eff_runtime)

    return (det, db, di, cg, mean_lum, std_lum,
        float(roi_mean), float(roi_bright), bool(bright_ok),
        float(white_frac_gray), float(white_frac_rgb))

def run_loop():
    """Main video capture loop - always runs for display, only logs when detection_active=True"""
    mon.worker = threading.current_thread()
    try:
        bars_h_roi = ph_bars_ref_roi(mon.bars_ref)
        int_h_list = ph_int_ref_list(mon.int_ref_paths) or [ph_int_ref(mon.int_ref)]
        ref_mean, ref_std, ref_bright = ref_roi_stats_from_paths(mon.int_ref_paths)
        # cache for UI grid annotations
        with mon.lock:
            mon._bars_h_roi = bars_h_roi
            mon._int_h_list = int_h_list
            mon._ref_mean = ref_mean
            mon._ref_bright = ref_bright
    except Exception as e:
        with mon.lock:
            mon.status = f"error: {e}"
            mon.running = False
            mon.video_running = False
        return

    # open source
    use_stream = isinstance(mon.stream_path, str) and mon.stream_path.strip() != ""
    backend = mon._resolve_backend()
    cap = None

    def _open_with_backend(_src, _backend):
        """Open camera and validate it stays consistent (no flickering between devices)"""
        if isinstance(_src, str) and not _src.isdigit():
            return cv2.VideoCapture(_src)  # URL or file path
        try:
            idx = int(_src)
        except Exception:
            idx = _src
        
        cap = cv2.VideoCapture(idx, _backend)
        
        # Validate camera stability: read frames and check they're similar
        # This prevents flickering between multiple cameras at the same index
        # macOS/OBS Virtual Camera needs more aggressive validation
        if cap and cap.isOpened():
            num_frames = 10 if platform.system() == "Darwin" else 5
            print(f"[Camera Validation] Testing stability of Source {idx} ({num_frames} frames)...")
            frame_means = []
            frame_shapes = []
            
            for i in range(num_frames):
                ret, frame = cap.read()
                if ret and frame is not None:
                    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    frame_means.append(float(np.mean(gray)))
                    frame_shapes.append(frame.shape)
                    # Small delay on macOS to help stabilize OBS Virtual Camera
                    if platform.system() == "Darwin" and i < num_frames - 1:
                        time.sleep(0.05)  # 50ms delay between frames
                else:
                    print(f"[Camera Validation] ✗ Failed to read frame {i+1}/{num_frames}")
                    return cap  # Return anyway, let caller handle
            
            # Check if frame properties are consistent
            if len(frame_means) >= 3:
                mean_variance = np.std(frame_means)
                shapes_match = len(set(frame_shapes)) == 1
                
                print(f"[Camera Validation] Mean brightness values: {[f'{m:.1f}' for m in frame_means]}")
                print(f"[Camera Validation] Std dev of means: {mean_variance:.1f}")
                print(f"[Camera Validation] Shapes consistent: {shapes_match}")
                
                # If variance is very high (>50), cameras might be switching
                if mean_variance > 50:
                    print(f"[Camera Validation] ⚠️  HIGH VARIANCE ({mean_variance:.1f}) - Camera may be unstable/flickering!")
                    print(f"[Camera Validation] This often means multiple cameras at same index")
                    print(f"[Camera Validation] Try a different Source index or Backend")
                elif not shapes_match:
                    print(f"[Camera Validation] ⚠️  Frame shapes changed - Camera unstable!")
                else:
                    print(f"[Camera Validation] ✓ Camera appears stable")
        
        return cap

    if use_stream:
        cap = cv2.VideoCapture(mon.stream_path.strip())
    else:
        # try selected backend, then fallback (Windows: MSMF<->DSHOW)
        cap = _open_with_backend(mon.src, backend)
        if (not cap) or (not cap.isOpened()):
            if platform.system() == "Windows":
                alt_backend = cv2.CAP_DSHOW if backend == cv2.CAP_MSMF else cv2.CAP_MSMF
                try:
                    if cap: cap.release()
                except Exception:
                    pass
                cap = _open_with_backend(mon.src, alt_backend)
            elif platform.system() == "Darwin":
                # Try several AVFoundation indices and pick the first non‑black feed.
                try:
                    if cap: cap.release()
                except Exception:
                    pass
                fourcc_sel = (mon.fourcc or "auto").upper()
                want_w, want_h = (1280, 720) if (mon.res_preset or "1080p").lower()=="720p" else (1920, 1080)
                # Check if OBS is running - if so, prioritize indices 1,2,3 over 0 (FaceTime)
                obs_running = False
                try:
                    result = subprocess.run(['pgrep', '-f', 'obs'], capture_output=True, timeout=2)
                    obs_running = result.returncode == 0
                except:
                    pass
                # If OBS is running, try indices 1,2,3 first (OBS Virtual Camera), then 0 (FaceTime)
                # Otherwise, try in normal order
                indices_to_try = [1, 2, 3, 0, 4, 5] if obs_running else [1, 0, 2, 3, 4, 5]
                print(f"[run_loop] Trying multiple AVFoundation indices to find a live source... (OBS running: {obs_running})")
                cap, used_idx = _mac_try_indices_for_nonblack(indices_to_try, cv2.CAP_AVFOUNDATION, want_w, want_h, fourcc_sel)
                if cap is None:
                    # last ditch: open the requested index even if black, so UI can still adjust
                    print(f"[run_loop] ⚠️  All indices returned black frames! Opening index {mon.src} anyway...")
                    try:
                        cap = cv2.VideoCapture(int(mon.src), cv2.CAP_AVFOUNDATION)
                        used_idx = int(mon.src)
                    except Exception:
                        cap = None
                        used_idx = None
                if used_idx is not None:
                    print(f"[run_loop] ✓ macOS using AVFoundation index {used_idx}")
    # Normalize capture on Windows/OBS to reduce flicker and colorspace issues
    try:
        cap.set(cv2.CAP_PROP_CONVERT_RGB, 1)
    except Exception:
        pass
    try:
        # Respect user-selected FOURCC, or leave as driver default when "auto"
        fc = (mon.fourcc or "auto").upper()
        if fc == "MJPG":
            cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'MJPG'))
        elif fc == "YUY2":
            cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'YUY2'))
        # else: auto -> do not force FOURCC
    except Exception:
        pass
    try:
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
    except Exception:
        pass
    try:
        cap.set(cv2.CAP_PROP_FPS, 30)
    except Exception:
        pass
    
    # CAMERA LOCKING: Prevent source switching/flickering
    # Windows: Lock camera handle to prevent enumeration issues
    # macOS: Lock OBS Virtual Camera to prevent flickering between devices
    try:
        if platform.system() == "Windows":
            # Force camera to stay locked by disabling auto-reconnect
            cap.set(cv2.CAP_PROP_SETTINGS, 0)  # Disable settings dialog
            # Read and discard a few frames to ensure camera is locked
            for _ in range(3):
                cap.read()
            print("[run_loop] ✓ Camera handle locked (Windows non-admin mode)")
        elif platform.system() == "Darwin":
            # macOS/AVFoundation: Lock camera by reading multiple frames and setting properties
            # This helps stabilize OBS Virtual Camera which can flicker between devices
            print("[run_loop] Locking camera handle (macOS)...")
            # Set buffer size to 1 to reduce latency and prevent frame queuing
            try:
                cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
            except Exception:
                pass
            # Read and discard multiple frames to ensure camera is locked
            # More frames for macOS to ensure OBS Virtual Camera stabilizes
            for i in range(10):
                ret, frame = cap.read()
                if ret and frame is not None:
                    # Verify we're getting consistent frames (not flickering)
                    if i == 0:
                        first_shape = frame.shape
                    elif i > 0 and frame.shape != first_shape:
                        print(f"[run_loop] ⚠️  Frame shape changed during lock: {first_shape} -> {frame.shape}")
                else:
                    print(f"[run_loop] ⚠️  Failed to read frame {i+1}/10 during lock")
            print("[run_loop] ✓ Camera handle locked (macOS)")
    except Exception as e:
        print(f"[run_loop] ⚠️  Camera locking failed: {e}")
        pass

    # Apply resolution preset (defaults to 1080p)
    try:
        rp = (mon.res_preset or "1080p").lower()
        if rp == "720p":
            w, h = 1280, 720
        else:
            w, h = 1920, 1080
        cap.set(cv2.CAP_PROP_FRAME_WIDTH,  w)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, h)
    except Exception:
        pass
    if not cap.isOpened():
        with mon.lock:
            mon.status = "error: cannot open source"
            mon.running = False
            mon.video_running = False
        print("[run_loop] ❌ ERROR: Cannot open video source!")
        print(f"[run_loop] Tried Source={mon.src}, Backend={mon.backend_name}")
        print("[run_loop] Try a different Source index or Backend")
        return
    
    # ensure we can read; bail early if not (retry once on Windows with alt backend)
    print("[run_loop] Testing if we can read frames from source...")
    ok_probe, probe_frame = cap.read()
    if ok_probe and probe_frame is not None:
        probe_gray = cv2.cvtColor(probe_frame, cv2.COLOR_BGR2GRAY)
        probe_mean = float(np.mean(probe_gray))
        print(f"[run_loop] ✓ Initial frame read successful: mean={probe_mean:.1f}, shape={probe_frame.shape}")
    else:
        print(f"[run_loop] ⚠️  Initial frame read failed: ok={ok_probe}, frame={'None' if probe_frame is None else 'valid'}")
    
    if (not ok_probe) and platform.system()=="Windows":
        try:
            cap.release()
        except Exception:
            pass
        alt_backend = cv2.CAP_DSHOW if backend == cv2.CAP_MSMF else cv2.CAP_MSMF
        cap = _open_with_backend(mon.src, alt_backend)
        ok_probe, _ = cap.read()
    if not ok_probe:
        with mon.lock:
            mon.status = "error: cannot read from source"
            mon.running = False
        cap.release()
        return

    # Warm-up: drop frames and give camera time to stabilize
    # Also establish baseline frame properties to detect camera switching
    print("[run_loop] Warming up capture device...")
    baseline_shape = None
    baseline_mean = None
    for i in range(20):
        ok, test_frame = cap.read()
        if ok and test_frame is not None:
            test_gray = cv2.cvtColor(test_frame, cv2.COLOR_BGR2GRAY)
            test_mean = float(np.mean(test_gray))
            if baseline_shape is None:
                baseline_shape = test_frame.shape
                baseline_mean = test_mean
            if i % 5 == 0:  # Log every 5th frame
                print(f"[run_loop] Warmup frame {i}: mean={test_mean:.1f}, shape={test_frame.shape}")
        time.sleep(0.05)  # 50ms between reads
    print(f"[run_loop] Warmup complete, baseline: shape={baseline_shape}, mean={baseline_mean:.1f}")

    # CSV file handle - only opened when detection starts
    csv_file = None
    csv_writer = None
    
    with mon.lock:
        mon.status = "video_ready"
        mon.video_running = True

    try:
        while True:
            with mon.lock:
                if not mon.video_running:
                    break
                # Open CSV file when detection first starts
                if mon.detection_active and csv_file is None:
                    csv_file, csv_writer = open_csv(mon.csv_path)
                    mon.test_start_time = datetime.now()
                    mon.status = "detecting"
                    print(f"[run_loop] Detection activated, logging to: {mon.csv_path}")
                cw, thr, st = mon.center_w, mon.thresh, mon.stable_frames
                dmean, dstd = mon.dark_mean, mon.dark_std
                # get anti-flicker tunables from monitor instance
                hold_ms = mon.hold_ms
                margin  = mon.margin

            ok, frame = cap.read()
            if not ok:
                time.sleep(0.05)
                continue
            
            # macOS: Monitor for camera switching/flickering
            # If frame properties change unexpectedly, the camera may have switched
            if platform.system() == "Darwin" and baseline_shape is not None:
                if frame.shape != baseline_shape:
                    print(f"[run_loop] ⚠️  CAMERA SWITCH DETECTED! Shape changed: {baseline_shape} -> {frame.shape}")
                    print(f"[run_loop] Attempting to re-lock camera...")
                    # Try to re-lock by reading multiple frames
                    for _ in range(5):
                        cap.read()
                        time.sleep(0.05)
                    # Update baseline
                    ok_check, frame_check = cap.read()
                    if ok_check and frame_check is not None:
                        baseline_shape = frame_check.shape
                        frame = frame_check
                        print(f"[run_loop] Camera re-locked, new baseline: {baseline_shape}")
                    else:
                        print(f"[run_loop] ⚠️  Failed to re-lock camera!")
                else:
                    # Check mean brightness - if it changes dramatically, camera may have switched
                    frame_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    frame_mean = float(np.mean(frame_gray))
                    if baseline_mean is not None and abs(frame_mean - baseline_mean) > 50:
                        # Mean changed significantly - might be camera switching
                        # But don't panic - just log it
                        if abs(frame_mean - baseline_mean) > 100:
                            print(f"[run_loop] ⚠️  Large brightness change: {baseline_mean:.1f} -> {frame_mean:.1f}")
                            # Update baseline to new value (camera may have stabilized on different source)
                            baseline_mean = frame_mean

            det, db, di, crop_img, mean_l, std_l, roi_mean, roi_bright, bright_ok, _wfg, _wfr = decide(
                frame, bars_h_roi, int_h_list, cw, thr, dmean, dstd, margin, ref_mean, ref_bright,
                white_frac_override=getattr(mon, "white_frac_gate", None),
                mean_gate_override=getattr(mon, "mean_gate", None)
            )

            with mon.lock:
                # update live artifacts
                mon.last_frame = frame
                # Throttle live-crop updates so UI refreshes don't perturb timing on slower systems
                now_sec = time.time()
                if mon.thumb_enabled and ((now_sec - mon._last_thumb_ts) * 1000.0 >= mon.thumb_every_ms):
                    # store a copy to decouple from OpenCV buffer reuse
                    mon.last_crop = crop_img.copy()
                    mon._last_thumb_ts = now_sec
                mon.last_metrics = {
                    "db": db, "di": di,
                    "mean": round(mean_l, 2), "std": round(std_l, 2),
                    "roi_mean": round(float(roi_mean), 1),
                    "roi_bright": round(float(roi_bright), 3),
                    "bright_ok": bool(bright_ok)
                }

                # stabilize
                if det == mon._raw_last:
                    mon._raw_stable += 1
                else:
                    # _raw_last stores the most recent raw detection result to track stability across frames.
                    mon._raw_last = det
                    mon._raw_stable = 1

                changed = False
                # Hysteresis: require the winning class to be clearly inside threshold by a margin
                det_ok = False
                if det == "BARS":
                    det_ok = (db < (thr + mon.margin))
                elif det == "INTERFACE":
                    # Accept either a close pHash OR a strong-bright ROI flag
                    det_ok = (di < (thr + mon.margin)) or bright_ok
                else:
                    det_ok = True  # OTHER/NO_SIGNAL has no phash gate

                now_ts = time.time()
                long_enough = (now_ts - mon._last_change_ts) * 1000.0 >= mon.hold_ms

                if mon._raw_stable >= st and det != mon.status and det_ok and long_enough:
                    prev = mon.status
                    mon.status = det
                    mon._last_change_ts = now_ts

                    # bump counters, track cycles, timestamps
                    now = datetime.now().isoformat(timespec="seconds")
                    if det == "BARS":
                        mon.count_bars += 1
                        mon.await_connect = True  # any non-connected period arms the cycle
                        print(f"[run_loop] State change: {prev} → BARS (db={db}, mean={mean_l:.1f})")
                    elif det == "INTERFACE":
                        mon.count_int  += 1
                        # Count a cycle when we reach INTERFACE after any non-connected period
                        if mon.await_connect:
                            mon.cycles += 1
                            mon.await_connect = False
                        print(f"[run_loop] State change: {prev} → INTERFACE (di={di}, roi_mean={roi_mean:.1f})")
                    elif det == "NO_SIGNAL":
                        mon.count_other += 1
                        mon.await_connect = True  # arm cycle when leaving connected state
                        print(f"[run_loop] State change: {prev} → NO_SIGNAL (mean={mean_l:.1f}, std={std_l:.1f}, thresholds: mean<{dmean}, std<{dstd})")
                    else:
                        mon.count_other += 1
                        mon.await_connect = True  # arm cycle when leaving connected state
                        print(f"[run_loop] State change: {prev} → {det}")

                    mon.last_seen[det] = now
                    mon._last = det

                    # log row
                    changed = True

            if not mon.video_running:
                break
            if not changed:
                time.sleep(0.05)
    finally:
        cap.release()
        if csv_file:
            csv_file.close()
        with mon.lock:
            if mon.video_running:
                mon.status = "stopped"
                mon.video_running = False
                mon.detection_active = False
                mon.running = False
            mon.worker = None

def _recalculate_elapsed_times_from_no_signal():
    """
    Post-process CSV to recalculate elapsed times:
    - Scope Connected: timestamp of Scope Connected - timestamp of previous No Signal
    - Scope Disconnected: timestamp of Scope Disconnected - timestamp of previous No Signal
    This is done per channel.
    """
    if not os.path.exists(mon.csv_path):
        print("[Post-Process] CSV file not found, skipping elapsed time recalculation")
        return
    
    try:
        # Read all rows from CSV, separating data from report text
        rows = []
        report_text = ""
        
        with open(mon.csv_path, 'r', newline='') as f:
            content = f.read()
        
        # Find where report starts (look for "=" * 80 pattern)
        report_start = content.find("=" * 80)
        if report_start >= 0:
            report_text = content[report_start:]
            content = content[:report_start]
        
        # Parse CSV data rows
        from io import StringIO
        f = StringIO(content)
        reader = csv.DictReader(f)
        for row in reader:
            # Skip rows that don't have required fields (might be empty or malformed)
            if not row.get('Video Channel') or not row.get('State') or not row.get('Timestamp'):
                continue
            rows.append(row)
        
        if not rows:
            print("[Post-Process] No data rows found in CSV")
            return
        
        # Track last "No Signal" timestamp per channel (separate tracking for each video channel 1-6)
        # This ensures elapsed times are calculated from the most recent "No Signal" for the SAME channel
        last_no_signal_ts = {}  # video_channel (int) -> timestamp (float, in seconds)
        
        # Process each row in chronological order and update elapsed times
        # Note: Rows should already be in chronological order from CSV, but we process sequentially
        for row in rows:
            video_channel = row.get('Video Channel', '').strip()
            state = row.get('State', '').strip()
            timestamp_str = row.get('Timestamp', '').strip()
            
            if not video_channel or not state or not timestamp_str:
                continue
            
            try:
                video_num = int(video_channel)
                # Parse timestamp and convert to seconds (consistent units for calculation)
                timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                if timestamp.tzinfo:
                    timestamp = timestamp.replace(tzinfo=None)
                timestamp_secs = timestamp.timestamp()  # Convert to seconds (float)
            except (ValueError, TypeError) as e:
                continue
            
            # Track "No Signal" timestamps per channel
            # Each channel maintains its own separate "last No Signal" timestamp
            if state == "No Signal":
                last_no_signal_ts[video_num] = timestamp_secs  # Update this channel's last No Signal time
                # Reset elapsed time for No Signal (not calculated from previous No Signal)
                row['Elapsed Secs'] = ''
            # Calculate elapsed time for Scope Connected/Disconnected
            # Uses the most recent "No Signal" timestamp for THIS SPECIFIC channel
            elif state in ["Scope Connected", "Scope Disconnected"]:
                if video_num in last_no_signal_ts:
                    # Calculate elapsed time: current timestamp - last No Signal timestamp for this channel
                    # Both timestamps are in seconds, so result is in seconds
                    elapsed = timestamp_secs - last_no_signal_ts[video_num]
                    row['Elapsed Secs'] = f"{elapsed:.2f}"
                else:
                    # No previous No Signal found for this channel, keep existing or set to empty
                    if not row.get('Elapsed Secs'):
                        row['Elapsed Secs'] = ''
        
        # Write updated rows back to CSV
        fieldnames = [
            "Timestamp", "Video Channel", "Console Serial", "Scope ID", "State",
            "Elapsed Secs", "full cycle count", "partial cycle count", "Disconnected Match Distance",
            "Connected Match Distance", "Event Type"
        ]
        
        # Write updated CSV with headers
        with open(mon.csv_path, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
        
        # Append report text back if it existed
        if report_text:
            with open(mon.csv_path, 'a', newline='') as f:
                f.write(report_text)
        
        # Also update Excel file if it exists
        if EXCEL_AVAILABLE:
            excel_path = _get_excel_path(mon.csv_path)
            if excel_path and os.path.exists(excel_path):
                try:
                    wb = load_workbook(excel_path)
                    ws = wb.active
                    
                    # Clear existing data (keep header row)
                    for row in range(2, ws.max_row + 1):
                        ws.delete_rows(row)
                    
                    # Write updated rows (convert numeric strings to numbers for Excel)
                    for row_data in rows:
                        # Convert numeric fields to proper types
                        elapsed_secs_str = row_data.get('Elapsed Secs', '').strip()
                        elapsed_secs = float(elapsed_secs_str) if elapsed_secs_str and elapsed_secs_str != '' else None
                        
                        cycle_num_str = str(row_data.get('full cycle count', '')).strip()
                        cycle_num = int(cycle_num_str) if cycle_num_str and cycle_num_str != '' and cycle_num_str.isdigit() else None
                        
                        partial_cycle_num_str = str(row_data.get('partial cycle count', '')).strip()
                        partial_cycle_num = int(partial_cycle_num_str) if partial_cycle_num_str and partial_cycle_num_str != '' and partial_cycle_num_str.isdigit() else None
                        
                        bars_dist_str = str(row_data.get('Disconnected Match Distance', '')).strip()
                        bars_dist = float(bars_dist_str) if bars_dist_str and bars_dist_str != '' else None
                        
                        int_dist_str = str(row_data.get('Connected Match Distance', '')).strip()
                        int_dist = float(int_dist_str) if int_dist_str and int_dist_str != '' else None
                        
                        video_channel_str = str(row_data.get('Video Channel', '')).strip()
                        video_channel = int(video_channel_str) if video_channel_str and video_channel_str.isdigit() else video_channel_str
                        
                        row_values = [
                            row_data.get('Timestamp', ''),
                            video_channel,
                            row_data.get('Console Serial', ''),
                            row_data.get('Scope ID', ''),
                            row_data.get('State', ''),
                            elapsed_secs,  # Number, not string
                            cycle_num,  # Number, not string
                            partial_cycle_num,  # Number, not string
                            bars_dist,  # Number, not string
                            int_dist,  # Number, not string
                            row_data.get('Event Type', '')
                        ]
                        ws.append(row_values)
                    
                    wb.save(excel_path)
                    print(f"[Post-Process] Updated Excel file: {excel_path}")
                except Exception as e:
                    print(f"[Post-Process] Error updating Excel file: {e}")
        
        print(f"[Post-Process] Recalculated elapsed times from No Signal for {len(rows)} rows")
        
    except Exception as e:
        print(f"[Post-Process] Error recalculating elapsed times: {e}")
        import traceback
        traceback.print_exc()

def _generate_and_append_report():
    """
    Generate a comprehensive test report with cycle statistics, timing analysis,
    and identification of incomplete cycles. Append to the CSV file.
    Returns: (report_text, csv_path)
    """
    report_lines = []
    report_lines.append("")
    report_lines.append("")
    report_lines.append("=" * 80)
    report_lines.append("BOOT CYCLE TEST REPORT")
    report_lines.append("=" * 80)
    report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append(f"Test Start: {mon.test_start_time.strftime('%Y-%m-%d %H:%M:%S') if mon.test_start_time else 'N/A'}")
    report_lines.append("")
    
    # Equipment Information
    report_lines.append("EQUIPMENT CONFIGURATION")
    report_lines.append("-" * 80)
    for console_key in ["console1", "console2", "console3"]:
        info = mon.equipment[console_key]
        console_num = console_key[-1]
        videos = ", ".join([f"Video {v}" for v in info["videos"]])
        report_lines.append(f"Console {console_num}: Serial={info['serial']}, Scope ID={info['scope_id']}, Channels=({videos})")
    report_lines.append("")
    
    # Detection Parameters
    report_lines.append("DETECTION PARAMETERS")
    report_lines.append("-" * 80)
    report_lines.append(f"Video Source: {getattr(mon, 'src', 'N/A')}")
    report_lines.append(f"Backend: {getattr(mon, 'backend', 'N/A')}")
    report_lines.append(f"FOURCC: {getattr(mon, 'fourcc', 'N/A')}")
    report_lines.append(f"Resolution: {getattr(mon, 'actual_w', '?')}x{getattr(mon, 'actual_h', '?')}")
    report_lines.append(f"Bars Threshold (pHash distance): {mon.thresh}")
    report_lines.append(f"Stable Frames Required: {mon.stable_frames}")
    roi_white_pct = getattr(mon, 'roi_white_pct', 0)
    roi_mean_gate = getattr(mon, 'roi_mean_gate', 0)
    roi_inset = getattr(mon, 'roi_inset', 0)
    report_lines.append(f"ROI White % Override: {roi_white_pct if roi_white_pct > 0 else 'Auto'}")
    report_lines.append(f"ROI Mean Gate Override: {roi_mean_gate if roi_mean_gate > 0 else 'Auto'}")
    report_lines.append(f"ROI Inset (px): {roi_inset}")
    report_lines.append("")
    
    # Per-Channel Statistics
    report_lines.append("PER-CHANNEL STATISTICS")
    report_lines.append("-" * 80)
    TILE_TO_VIDEO = [1, 3, 5, 2, 4, 6]
    
    for tile_idx in range(GRID_FEEDS):
        video_num = TILE_TO_VIDEO[tile_idx]
        console_serial, scope_id = _get_equipment_for_video(video_num)
        
        report_lines.append(f"\nVideo Channel {video_num} (Console: {console_serial}, Scope: {scope_id})")
        report_lines.append(f"  Complete Cycles: {mon.tile_complete_cycles[tile_idx]}")
        report_lines.append(f"  Total Transitions to Connected: {mon.tile_counts[tile_idx]}")
        report_lines.append(f"  Total Transitions to Disconnected: {mon.tile_disconnected_counts[tile_idx]}")
        
        # Parse CSV to get detailed statistics about Scope Disconnected states
        disconnected_data = []
        try:
            if os.path.exists(mon.csv_path):
                with open(mon.csv_path, 'r', newline='') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if (row.get('Video Channel') == str(video_num) and 
                            row.get('State') == 'Scope Disconnected'):
                            elapsed = row.get('Elapsed Secs', '')
                            cycle = row.get('full cycle count', '')
                            if elapsed:
                                try:
                                    disconnected_data.append({
                                        'elapsed': float(elapsed),
                                        'cycle': int(cycle) if cycle else 0
                                    })
                                except (ValueError, TypeError):
                                    pass
        except Exception as e:
            print(f"[Report] Error parsing CSV for disconnected stats: {e}")
        
        if disconnected_data:
            report_lines.append(f"  Scope Disconnected Events:")
            report_lines.append(f"    Total Count: {len(disconnected_data)}")
            elapsed_times = [d['elapsed'] for d in disconnected_data]
            if elapsed_times:
                report_lines.append(f"    Elapsed Times from Test Start:")
                report_lines.append(f"      First: {min(elapsed_times):.2f} seconds")
                report_lines.append(f"      Last: {max(elapsed_times):.2f} seconds")
                report_lines.append(f"      Average: {sum(elapsed_times) / len(elapsed_times):.2f} seconds")
            cycle_numbers = [d['cycle'] for d in disconnected_data if d['cycle'] is not None]
            if cycle_numbers:
                unique_cycles = sorted(set(cycle_numbers))
                report_lines.append(f"    Cycle Numbers: {', '.join(map(str, unique_cycles))}")
        else:
            report_lines.append(f"  No Scope Disconnected events recorded")
        
        cycle_times = mon.tile_cycle_times[tile_idx]
        if cycle_times:
            avg_time = sum(cycle_times) / len(cycle_times)
            min_time = min(cycle_times)
            max_time = max(cycle_times)
            report_lines.append(f"  Reconnection Times (Disconnected→Connected):")
            report_lines.append(f"    Average: {avg_time:.2f} seconds")
            report_lines.append(f"    Min: {min_time:.2f} seconds")
            report_lines.append(f"    Max: {max_time:.2f} seconds")
            report_lines.append(f"    Count: {len(cycle_times)}")
            
            # Identify timing regularities (standard deviation, outliers)
            if len(cycle_times) > 1:
                import statistics
                std_dev = statistics.stdev(cycle_times)
                report_lines.append(f"    Std Deviation: {std_dev:.2f} seconds")
                
                # Flag outliers (> 2 standard deviations from mean)
                outliers = [t for t in cycle_times if abs(t - avg_time) > 2 * std_dev]
                if outliers:
                    report_lines.append(f"    ⚠ Outliers detected: {len(outliers)} reconnections took unusually long/short")
        else:
            report_lines.append(f"  No reconnection timing data recorded")
        
        # Check for incomplete cycles (stuck in Disconnected state)
        if mon.tile_disconnected_start[tile_idx] is not None:
            report_lines.append(f"  ⚠ WARNING: Channel ended in Scope Disconnected state (incomplete cycle)")
    
    report_lines.append("")
    report_lines.append("=" * 80)
    report_lines.append("INCOMPLETE CYCLES ANALYSIS")
    report_lines.append("-" * 80)
    
    # Analyze state history for incomplete cycles
    incomplete_found = False
    for tile_idx in range(GRID_FEEDS):
        video_num = TILE_TO_VIDEO[tile_idx]
        history = mon.tile_state_history[tile_idx]
        
        # Look for sequences that started Disconnected but didn't reach Connected
        for i in range(len(history) - 1):
            if history[i]["state"] == "BARS":
                # Check if next state is NOT INTERFACE
                if i + 1 < len(history) and history[i + 1]["state"] != "INTERFACE":
                    incomplete_found = True
                    ts = datetime.fromtimestamp(history[i]["timestamp"]).strftime('%H:%M:%S')
                    next_state = history[i + 1]["state"]
                    report_lines.append(f"Video {video_num}: Disconnected at {ts} → {label_for(next_state)} (did not reconnect)")
        
        # Check if ended in Disconnected
        if history and history[-1]["state"] == "BARS":
            incomplete_found = True
            ts = datetime.fromtimestamp(history[-1]["timestamp"]).strftime('%H:%M:%S')
            report_lines.append(f"Video {video_num}: Disconnected at {ts} → Test ended (incomplete)")
    
    if not incomplete_found:
        report_lines.append("No incomplete cycles detected.")
    
    report_lines.append("")
    report_lines.append("=" * 80)
    report_lines.append("END OF REPORT")
    report_lines.append("=" * 80)
    
    report_text = "\n".join(report_lines)
    
    # Append report to CSV file
    try:
        with open(mon.csv_path, "a", newline="") as f:
            # Write report as comments/text at end of CSV
            f.write("\n")
            f.write(report_text)
            f.write("\n")
        print(f"[Report] Appended to CSV: {mon.csv_path}")
    except Exception as e:
        print(f"[Report] Error appending to CSV: {e}")
    
    return report_text, mon.csv_path

def _reveal_in_file_browser(file_path):
    """Open the system file browser and reveal the specified file (cross-platform)."""
    try:
        abs_path = os.path.abspath(file_path)
        if not os.path.exists(abs_path):
            print(f"[Reveal] File not found: {abs_path}")
            return False
        
        if platform.system() == "Darwin":  # macOS
            subprocess.run(["open", "-R", abs_path], check=False)
        elif platform.system() == "Windows":
            subprocess.run(["explorer", "/select,", abs_path], check=False)
        else:  # Linux
            # Try to open the containing directory
            folder = os.path.dirname(abs_path)
            subprocess.run(["xdg-open", folder], check=False)
        return True
    except Exception as e:
        print(f"[Reveal] Error: {e}")
        return False

# ---------- web app ----------
app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)

@app.get("/")
def index():
    # Enhanced UI with pre-start modal for equipment identification
    return """<!doctype html><html><head><meta charset='utf-8'><title>Boot Cycle Logger 6‑ch</title>
    <style>
    :root{
      --bg:#0f172a; --panel:#0b1220; --border:#1f2937; --text:#e5e7eb; --muted:#94a3b8;
      --ok:#16a34a; --bad:#dc2626; --nosig:#6b7280; --other:#0ea5e9; --btn:#38bdf8; --btn2:#1f2937;
    }
    html,body{height:100%}
    body{margin:0;background:var(--bg);color:var(--text);font-family:system-ui,-apple-system,Segoe UI,Roboto,sans-serif}
    .wrap{display:grid;grid-template-columns: 2fr 1fr; min-height:100vh;}
    .left{padding:16px;background:var(--panel);border-left:1px solid var(--border); min-width:240px; max-width:420px;}
    .right{display:flex;flex-direction:column;gap:18px;padding:24px 18px 18px 36px;justify-content:flex-start;}
    h1{margin:4px 0 12px 0;font-size:20px;color:#38bdf8}
    h2{margin:16px 0 8px 0;font-size:16px;color:#38bdf8}
    label{display:block;font-size:12px;color:var(--muted);margin-bottom:6px}
    .row{margin-bottom:12px}
    button{background:var(--btn);color:#0b1220;font-weight:700;border:none;padding:10px 14px;border-radius:8px;cursor:pointer;transition:all 0.15s ease}
    button:hover{background:#0ea5e9;transform:translateY(-1px);box-shadow:0 4px 12px rgba(14,165,233,0.3)}
    button:active{transform:translateY(0);box-shadow:0 2px 4px rgba(14,165,233,0.2)}
    button.secondary{background:var(--btn2);color:var(--text)}
    button.secondary:hover{background:#1e293b;border:1px solid var(--btn);transform:translateY(-1px);box-shadow:0 4px 12px rgba(148,163,184,0.2)}
    button.secondary:active{transform:translateY(0);box-shadow:0 2px 4px rgba(148,163,184,0.15)}
    button:disabled{opacity:0.5;cursor:not-allowed}
    input,select{width:100%;padding:8px;border-radius:6px;background:var(--panel);color:var(--text);border:1px solid var(--border)}
    img.source{width:100%;max-width:100%;height:auto;display:block;border:1px solid var(--border);border-radius:10px;background:#000;aspect-ratio:16/9;object-fit:contain}
    .pillgrid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));grid-template-rows:repeat(2,1fr);gap:12px;margin-top:16px;width:100%}
    .pill{display:flex;align-items:center;justify-content:center;height:44px;border-radius:999px;font-weight:800;border:1px solid #00000022;box-shadow:0 6px 20px rgba(0,0,0,.25), inset 0 1px 0 rgba(255,255,255,.08);min-width:0}
    .ok{background:var(--ok)} .bad{background:var(--bad)} .nosig{background:var(--nosig)} .other{background:var(--other)}
    .pill span{color:white;white-space:nowrap;text-overflow:ellipsis;overflow:hidden;font-size:11px;}
    pre{font:12px ui-monospace,Menlo,Consolas,monospace;color:var(--muted);white-space:pre-wrap}
    
    /* Modal styles */
    .modal{display:none;position:fixed;z-index:1000;left:0;top:0;width:100%;height:100%;background:rgba(0,0,0,0.8);align-items:center;justify-content:center}
    .modal.show{display:flex}
    .modal-content{background:var(--panel);border:1px solid var(--border);border-radius:12px;padding:32px;max-width:600px;width:90%;max-height:90vh;overflow-y:auto}
    .modal-content h2{margin-top:0}
    .console-group{background:var(--bg);padding:16px;border-radius:8px;margin-bottom:16px;border:1px solid var(--border)}
    .console-group label{font-size:13px;font-weight:600;color:var(--text)}
    .console-group input{margin-bottom:8px}
    .equipment-info{background:var(--bg);padding:12px;border-radius:6px;margin-bottom:12px;font-size:11px;line-height:1.6}
    .report-view{background:var(--bg);padding:16px;border-radius:8px;margin-top:16px;max-height:400px;overflow-y:auto;display:none}
    .report-view.show{display:block}
    
    /* Test Running Indicator */
    .test-indicator{display:inline-block !important;width:48px;height:48px;border-radius:50%;margin-left:12px;vertical-align:middle;opacity:0;transition:opacity 0.3s ease;position:relative;top:2px;background:transparent}
    .test-indicator.running{background:#ef4444 !important;opacity:1 !important;animation:pulse 1.5s ease-in-out infinite}
    @keyframes pulse{0%,100%{opacity:1 !important;box-shadow:0 0 0 0 rgba(239,68,68,0.7)}50%{opacity:0.6 !important;box-shadow:0 0 0 16px rgba(239,68,68,0)}}
    
    @media (max-width: 900px){.wrap{grid-template-columns:1fr}.left{max-width:none;border-left:none;border-top:1px solid var(--border)}.right{padding:14px}}
    </style></head><body>
    
    <!-- Equipment Setup Modal -->
    <div id='equipModal' class='modal show'>
      <div class='modal-content'>
        <h2>Equipment Setup</h2>
        <p style='color:var(--muted);margin-bottom:24px'>Enter console serial numbers and scope IDs before starting the test.</p>
        
        <div style='display:grid;grid-template-columns:1fr 1fr;gap:20px'>
          <div class='console-group'>
            <h3 style='margin-top:0;color:#38bdf8;font-size:14px'>Console 1 (Videos 1 & 2)</h3>
            <label>Console Serial Number</label>
            <input id='console1_serial' placeholder='e.g., SN123456' value=''>
            <label>Scope ID</label>
            <input id='console1_scope' placeholder='e.g., SCOPE-A1' value=''>
          </div>
          
          <div class='console-group'>
            <h3 style='margin-top:0;color:#38bdf8;font-size:14px'>Console 2 (Videos 3 & 4)</h3>
            <label>Console Serial Number</label>
            <input id='console2_serial' placeholder='e.g., SN123457' value=''>
            <label>Scope ID</label>
            <input id='console2_scope' placeholder='e.g., SCOPE-B2' value=''>
          </div>
          
          <div class='console-group' style='grid-column:1/-1'>
            <h3 style='margin-top:0;color:#38bdf8;font-size:14px'>Console 3 (Videos 5 & 6)</h3>
            <label>Console Serial Number</label>
            <input id='console3_serial' placeholder='e.g., SN123458' value=''>
            <label>Scope ID</label>
            <input id='console3_scope' placeholder='e.g., SCOPE-C3' value=''>
          </div>
        </div>
        
        <button onclick='saveEquipmentAndContinue()' style='width:100%;padding:14px'>Save and Continue</button>
      </div>
    </div>
    
    <div class='wrap'>
      <div class='right'>
        <img id='grid' class='source' src='/thumb'>
        <div class='pillgrid'>
          <div id='p1' class='pill'><span>VIDEO 1</span></div>
          <div id='p3' class='pill'><span>VIDEO 3</span></div>
          <div id='p5' class='pill'><span>VIDEO 5</span></div>
          <div id='p2' class='pill'><span>VIDEO 2</span></div>
          <div id='p4' class='pill'><span>VIDEO 4</span></div>
          <div id='p6' class='pill'><span>VIDEO 6</span></div>    
        </div>
      </div>
      <div class='left'>
        <h1>Boot Cycle Logger <span id='testIndicator' class='test-indicator'></span></h1>
        
        <div class='equipment-info' id='equipInfo' style='display:none'>
          <strong>Equipment Configuration:</strong><br>
          <span id='equipSummary'></span>
        </div>
        
        <div style='display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px'>
          <button onclick='startTest()' id='startBtn'>Start Test</button>
          <button class='secondary' onclick='endTest()' id='endBtn'>End Test</button>
          <button class='secondary' onclick='showEquipModal()'>Edit Equipment</button>
          <button class='secondary' onclick='probe()'>Probe source</button>
          <button class='secondary' onclick='resetTallies()'>Reset tallies</button>
        </div>
        
        <div class='timer-section' style='background:var(--bg);padding:12px;border-radius:8px;margin-bottom:12px;border:1px solid var(--border)'>
          <div style='display:flex;flex-direction:column;gap:12px'>
            <div style='display:flex;align-items:center;gap:12px;flex-wrap:wrap'>
              <label style='display:flex;align-items:center;gap:6px;cursor:pointer'>
                <input type='checkbox' id='timerEnabled' style='cursor:pointer'>
                <span style='color:#38bdf8;font-size:14px;font-weight:500'>Testing Timer:</span>
              </label>
              <div style='display:flex;align-items:center;gap:8px'>
                <input type='number' id='timerHours' value='0' min='0' max='23' style='width:60px;padding:4px;border:1px solid var(--border);border-radius:4px;background:var(--bg2);color:var(--text)' disabled>
                <span style='color:var(--text2)'>hrs</span>
                <input type='number' id='timerMinutes' value='0' min='0' max='59' style='width:60px;padding:4px;border:1px solid var(--border);border-radius:4px;background:var(--bg2);color:var(--text)' disabled>
                <span style='color:var(--text2)'>min</span>
                <input type='number' id='timerSeconds' value='0' min='0' max='59' style='width:60px;padding:4px;border:1px solid var(--border);border-radius:4px;background:var(--bg2);color:var(--text)' disabled>
                <span style='color:var(--text2)'>sec</span>
              </div>
            </div>
            <div id='timerDisplay' style='display:block;font-weight:bold;color:#ef4444;font-size:20px;min-width:150px;padding:8px 16px;background:var(--bg2);border-radius:4px;border:2px solid #ef4444;text-align:center;margin-top:4px'>
              <span id='timerTime'>00:00:00</span>
            </div>
          </div>
        </div>
        
        <div class='camera-section' style='background:var(--bg);padding:12px;border-radius:8px;margin-bottom:12px;border:1px solid var(--border)'>
          <h3 style='margin-top:0;color:#38bdf8;font-size:14px;margin-bottom:12px'>Camera Source</h3>
          <div style='display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px'>
            <div class='row'><label>Source</label><input id='src' value='""" + str(mon.src) + """'></div>
            <div class='row'><label>Backend</label><select id='backend'><option>auto</option><option>MSMF</option><option>DSHOW</option><option>AVFOUNDATION</option><option>V4L2</option></select></div>
            <div class='row'><label>FOURCC</label><select id='fourcc'><option>auto</option><option>MJPG</option><option>YUY2</option></select></div>
            <div class='row'><label>Resolution</label><select id='res'><option>1080p</option><option>720p</option></select></div>
          </div>
        <div style='display:flex;gap:8px;margin-top:8px'>
          <button onclick='connectCamera()' style='flex:1'>Connect to Camera</button>
          <button class='secondary' onclick='autoDetectCamera()' style='flex:1'>Auto-Detect OBS</button>
          <button class='secondary' onclick='listAvailableCameras()' style='flex:1'>List Cameras</button>
        </div>
        <div style='display:flex;gap:8px;margin-top:8px'>
          <button class='secondary' onclick='resetCameraSystem()' style='flex:1'>Reset Camera System</button>
          <button class='secondary' onclick='getSystemCameras()' style='flex:1'>System Camera Report</button>
        </div>
        <div style='display:flex;gap:8px;margin-top:8px'>
          <button class='secondary' onclick='forceRestartApp()' style='flex:1;background:#dc2626;color:#fecaca'>Force Restart App</button>
          <button class='secondary' onclick='showRestartInstructions()' style='flex:1;background:#1e293b;color:#e2e8f0'>Manual Restart</button>
        </div>
          <div id='cameraStatus' style='margin-top:8px;font-size:11px;padding:6px;border-radius:4px;display:none'></div>
          <div id='cameraList' style='margin-top:8px;font-size:10px;display:none'></div>
        </div>
        
        <h3 style='margin-top:16px;color:#38bdf8;font-size:14px;margin-bottom:12px'>Detection Settings</h3>
        <div class='row'><label>Bars Threshold</label><input id='thr' value='""" + str(mon.thresh) + """'></div>
        <div class='row'><label>Stable Frames</label><input id='st' value='""" + str(mon.stable_frames) + """'></div>
        <div class='row'><label>ROI white % (0-100, overrides auto)</label><input id='white_pct' value='""" + (str(mon.white_frac_gate) if mon.white_frac_gate is not None else '') + """'></div>
        <div class='row'><label>ROI mean gate (0-255, overrides auto)</label><input id='mean_gate' value='""" + (str(mon.mean_gate) if mon.mean_gate is not None else '') + """'></div>
        <div class='row'><label>ROI inset px (shrink ROI)</label><input id='roi_inset' value='""" + str(getattr(mon, 'roi_inset_px', 1)) + """'></div>
        <div class='row' style='display:flex;gap:8px;flex-wrap:wrap'>
          <div id='csvInfo' style='font-size:12px;color:#94a3b8;margin-top:6px;width:100%'>CSV: <code id='csvName'>-</code></div>
        </div>
        <pre id='probeOut'></pre>
        
        <!-- Test Report Modal -->
        <div id='reportModal' class='modal' style='display:none'>
          <div class='modal-content' style='max-width:800px;max-height:80vh;overflow-y:auto'>
            <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:20px'>
              <h2 style='margin:0'>Test Report</h2>
              <button onclick='closeReportModal()' style='background:none;border:none;color:var(--muted);font-size:24px;cursor:pointer;padding:0;width:30px;height:30px;display:flex;align-items:center;justify-content:center'>&times;</button>
            </div>
            <pre id='reportText' style='font-size:12px;line-height:1.5;background:var(--bg-secondary);padding:16px;border-radius:8px;overflow-x:auto;white-space:pre-wrap;word-wrap:break-word'></pre>
            <div style='display:flex;gap:12px;margin-top:20px'>
              <button onclick='goToCSV()' id='goToCsvBtn' style='flex:1;padding:12px'>Go to CSV/Excel</button>
              <button onclick='closeReportModal()' class='secondary' style='flex:1;padding:12px'>Close</button>
            </div>
          </div>
        </div>
      </div>
    </div>
    <script>
  // Mapping: tile index (0..5) -> VIDEO # and pill id (1,3,5,2,4,6)
  const TILE_TO_PILL = [1, 3, 5, 2, 4, 6]; // tile index -> VIDEO# / pill id mapping

  function bust(u){ return u + (u.includes('?') ? '&' : '?') + 't=' + Date.now(); }

  async function startVideo(){
    // Start video feed only (no detection)
    await fetch('/start_video', { method: 'POST' });
  }

  // Timer state
  let timerInterval = null;
  let timerEndTime = null;

  // Enable/disable timer inputs based on checkbox
  document.addEventListener('DOMContentLoaded', function() {
    const timerEnabled = document.getElementById('timerEnabled');
    const timerHours = document.getElementById('timerHours');
    const timerMinutes = document.getElementById('timerMinutes');
    const timerSeconds = document.getElementById('timerSeconds');
    
    if (timerEnabled) {
      timerEnabled.addEventListener('change', function() {
        const enabled = this.checked;
        timerHours.disabled = !enabled;
        timerMinutes.disabled = !enabled;
        timerSeconds.disabled = !enabled;
      });
    }
  });

  // Timer countdown function
  function updateTimer() {
    const timerTimeEl = document.getElementById('timerTime');
    if (!timerTimeEl) return;
    
    if (!timerEndTime) {
      timerTimeEl.textContent = '00:00:00';
      return;
    }
    
    const now = Date.now();
    const remaining = Math.max(0, timerEndTime - now);
    
    if (remaining <= 0) {
      // Timer reached zero - auto-stop test
      clearInterval(timerInterval);
      timerInterval = null;
      timerEndTime = null;
      timerTimeEl.textContent = '00:00:00';
      console.log('Timer expired - auto-stopping test...');
      endTest();
      return;
    }
    
    const hours = Math.floor(remaining / 3600000);
    const minutes = Math.floor((remaining % 3600000) / 60000);
    const seconds = Math.floor((remaining % 60000) / 1000);
    
    const timeStr = String(hours).padStart(2, '0') + ':' + 
                   String(minutes).padStart(2, '0') + ':' + 
                   String(seconds).padStart(2, '0');
    timerTimeEl.textContent = timeStr;
  }

  async function startTest(){
    try {
      console.log('Start Test button clicked');
      // Stop any existing timer
      if (timerInterval) {
        clearInterval(timerInterval);
        timerInterval = null;
      }
      timerEndTime = null;
      const timerTimeEl = document.getElementById('timerTime');
      if (timerTimeEl) timerTimeEl.textContent = '00:00:00';
      
      // Start detection/logging (video should already be running)
      console.log('Calling /start_detection...');
      
      // Add timeout to prevent hanging
      const controller = new AbortController();
      const timeoutId = setTimeout(() => controller.abort(), 10000); // 10 second timeout
      
      let r;
      try {
        r = await fetch('/start_detection', { 
          method: 'POST',
          signal: controller.signal
        });
        clearTimeout(timeoutId);
        console.log('Response status:', r.status, r.statusText);
      } catch (fetchError) {
        clearTimeout(timeoutId);
        if (fetchError.name === 'AbortError') {
          console.error('Request timed out after 10 seconds');
          alert('Error: Request timed out. The server may be busy. Please try again.');
        } else {
          console.error('Fetch error:', fetchError);
          alert('Error connecting to server: ' + fetchError.message);
        }
        return;
      }
      
      if (!r.ok) {
        const errorText = await r.text();
        console.error('HTTP error:', r.status, errorText);
        alert('Error starting test: ' + (errorText || r.statusText));
        return;
      }
      
      const j = await r.json();
      console.log('Response JSON:', j);
      
      if (j.ok) {
        console.log('Test started successfully, CSV:', j.csv_path);
        // Show running indicator immediately - use multiple methods to ensure it works
        const indicator = document.getElementById('testIndicator');
        if (indicator) {
          // Add the running class
          indicator.classList.add('running');
          // Force visibility with multiple style properties
          indicator.style.setProperty('opacity', '1', 'important');
          indicator.style.setProperty('display', 'inline-block', 'important');
          indicator.style.setProperty('background', '#ef4444', 'important');
          indicator.style.setProperty('visibility', 'visible', 'important');
          console.log('Indicator updated - should be visible now');
          console.log('Indicator classes:', indicator.className);
          console.log('Indicator computed style:', window.getComputedStyle(indicator).opacity);
        } else {
          console.error('testIndicator element not found!');
          alert('ERROR: testIndicator element not found in DOM!');
        }
        // Hide report modal if previously shown
        const reportModal = document.getElementById('reportModal');
        if (reportModal) reportModal.style.display = 'none';
        
        // Start timer if enabled
        const timerEnabled = document.getElementById('timerEnabled');
        if (timerEnabled && timerEnabled.checked) {
          const hours = parseInt(document.getElementById('timerHours').value) || 0;
          const minutes = parseInt(document.getElementById('timerMinutes').value) || 0;
          const seconds = parseInt(document.getElementById('timerSeconds').value) || 0;
          const totalMs = (hours * 3600 + minutes * 60 + seconds) * 1000;
          
          console.log(`Timer values: ${hours}h ${minutes}m ${seconds}s = ${totalMs}ms`);
          
          if (totalMs > 0) {
            timerEndTime = Date.now() + totalMs;
            updateTimer(); // Update immediately
            timerInterval = setInterval(updateTimer, 1000); // Update every second
            console.log(`Timer started: ${hours}h ${minutes}m ${seconds}s`);
          } else {
            // Reset to 00:00:00 if timer is 0
            if (timerTimeEl) timerTimeEl.textContent = '00:00:00';
          }
        }
      } else {
        console.error('Start test failed:', j.error || 'Unknown error');
        alert('Failed to start test: ' + (j.error || 'Unknown error'));
      }
    } catch(e) {
      console.error('Error starting test:', e);
      alert('Error starting test: ' + e.message);
    }
  }

  // Debounce helper
  function debounce(fn, ms){
    let t; return (...args)=>{ clearTimeout(t); t=setTimeout(()=>fn(...args), ms); };
  }

  // Push live threshold adjustments
  async function pushAdjust(){
    const body = {
      white_pct: (function(v){ if(v===''||v==null) return null; v=parseFloat(v); return isNaN(v)?null:v; })(document.getElementById('white_pct').value),
      mean_gate: (function(v){ if(v===''||v==null) return null; v=parseFloat(v); return isNaN(v)?null:v; })(document.getElementById('mean_gate').value),
      roi_inset: (function(v){ if(v===''||v==null) return null; v=parseInt(v,10); return isNaN(v)?null:v; })(document.getElementById('roi_inset').value)
    };
    await fetch('/adjust', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify(body)
    });
  }

  const debouncedAdjust = debounce(pushAdjust, 250);
  document.getElementById('white_pct').addEventListener('input', debouncedAdjust);
  document.getElementById('mean_gate').addEventListener('input', debouncedAdjust);
  document.getElementById('roi_inset').addEventListener('input', debouncedAdjust);

  async function endTest(){
    try {
      console.log('End Test button clicked');
      // Stop timer if running
      if (timerInterval) {
        clearInterval(timerInterval);
        timerInterval = null;
      }
      timerEndTime = null;
      const timerTimeEl = document.getElementById('timerTime');
      if (timerTimeEl) timerTimeEl.textContent = '00:00:00';
      
      // Stop running indicator immediately
      const indicator = document.getElementById('testIndicator');
      if (indicator) {
        indicator.classList.remove('running');
        indicator.style.opacity = '0';
      }
      
      console.log('Calling /end_test endpoint...');
      const r = await fetch('/end_test', {method: 'POST'});
      
      console.log('End test response status:', r.status, r.statusText);
      
      if (!r.ok) {
        const errorText = await r.text();
        console.error('HTTP error:', r.status, errorText);
        alert('Error ending test: ' + (errorText || r.statusText));
        return;
      }
      
      const j = await r.json();
      console.log('End test response:', j);
      
      if (j.report) {
        // Display the report in modal
        const reportTextEl = document.getElementById('reportText');
        const reportModalEl = document.getElementById('reportModal');
        
        if (reportTextEl && reportModalEl) {
          reportTextEl.textContent = j.report;
          reportModalEl.style.display = 'block';
          
          console.log('Report displayed in modal, length:', j.report.length);
        } else {
          console.error('Report elements not found in DOM');
        }
      } else {
        console.warn('No report in response. Response:', j);
        if (j.error) {
          alert('Error generating report: ' + j.error);
        }
      }
    } catch(e) {
      console.error('Error ending test:', e);
      alert('Exception ending test: ' + e.message);
      // Still stop indicator on error
      document.getElementById('testIndicator').classList.remove('running');
    }
  }

  async function goToCSV(){
    try {
      await fetch('/reveal_csv', {method: 'POST'});
    } catch(e) {
      console.error('Error opening CSV folder:', e);
    }
  }

  function closeReportModal(){
    document.getElementById('reportModal').style.display = 'none';
  }

  // Force OBS detection with detailed logging
  async function forceOBSDetection(){
    const statusEl = document.getElementById('cameraStatus');
    statusEl.style.display = 'block';
    statusEl.style.background = '#1e293b';
    statusEl.style.color = '#94a3b8';
    statusEl.textContent = '🔍 Force detecting OBS Virtual Camera...';
    
    try {
      // Try different OBS names manually
      const obsNames = [
        'OBS Virtual Camera',
        'OBS-Camera', 
        'OBS Virtual Source',
        'OBS Virtual Camera (DirectShow)',
        'OBS Virtual Camera (Media Foundation)'
      ];
      
      let found = false;
      
      for (const name of obsNames) {
        statusEl.textContent = `🔍 Trying: ${name}...`;
        
        try {
          // Update source field and try to connect
          document.getElementById('src').value = name;
          
          const settings = {
            src: name,
            backend: 'DSHOW',
            fourcc: 'auto',
            res: '1080p'
          };
          
          await fetch('/update_camera_settings', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify(settings)
          });
          
          const r = await fetch('/start_video', {method: 'POST'});
          const j = await r.json();
          
          if (j.ok) {
            statusEl.style.background = '#064e3b';
            statusEl.style.color = '#10b981';
            statusEl.textContent = `✓ Found OBS Virtual Camera: ${name}`;
            found = true;
            break;
          }
        } catch(e) {
          console.log(`Failed to connect to ${name}:`, e);
          continue;
        }
      }
      
      if (!found) {
        statusEl.style.background = '#7f1d1d';
        statusEl.style.color = '#ef4444';
        statusEl.textContent = '✗ OBS Virtual Camera not found. Make sure Virtual Camera is started in OBS.';
      }
      
    } catch(e) {
      statusEl.style.background = '#7f1d1d';
      statusEl.style.color = '#ef4444';
      statusEl.textContent = '✗ Error: ' + e.message;
    }
  }

  // Connect to camera with current settings
  async function connectCamera(){
    const statusEl = document.getElementById('cameraStatus');
    statusEl.style.display = 'block';
    statusEl.style.background = '#1e293b';
    statusEl.style.color = '#94a3b8';
    statusEl.textContent = '⏳ Connecting to camera...';
    
    try {
      // First, update the camera settings on the server
      const settings = {
        src: document.getElementById('src').value,
        backend: document.getElementById('backend').value,
        fourcc: document.getElementById('fourcc').value,
        res: document.getElementById('res').value
      };
      
      // Update monitor settings
      await fetch('/update_camera_settings', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify(settings)
      });
      
      // Then restart the video feed
      const r = await fetch('/start_video', {method: 'POST'});
      const j = await r.json();
      
      if (j.ok) {
        statusEl.style.background = '#064e3b';
        statusEl.style.color = '#10b981';
        statusEl.textContent = `✓ Connected to Source ${settings.src} (${settings.backend})`;
        
        setTimeout(() => {
          statusEl.style.display = 'none';
        }, 3000);
      } else {
        statusEl.style.background = '#7f1d1d';
        statusEl.style.color = '#ef4444';
        statusEl.textContent = '✗ Failed to connect: ' + (j.error || 'Unknown error');
      }
    } catch(e) {
      statusEl.style.background = '#7f1d1d';
      statusEl.style.color = '#ef4444';
      statusEl.textContent = '✗ Error: ' + e.message;
    }
  }

  // Auto-detect camera
  async function autoDetectCamera(){
    const statusEl = document.getElementById('cameraStatus');
    statusEl.style.display = 'block';
    statusEl.style.background = '#1e293b';
    statusEl.style.color = '#94a3b8';
    statusEl.textContent = '🔍 Auto-detecting camera (OBS Virtual Camera or USB Capture)...';
    
    try {
      const r = await fetch('/auto_detect_camera');
      if (!r.ok) {
        throw new Error(`HTTP ${r.status}: ${r.statusText}`);
      }
      const j = await r.json();
      
      if (j.ok) {
        // Update UI with detected settings
        document.getElementById('src').value = j.index;
        // Set backend dropdown
        const backendSelect = document.getElementById('backend');
        if (backendSelect) {
          const backendLower = (j.backend || '').toLowerCase();
          for (let opt of backendSelect.options) {
            if (opt.value.toLowerCase() === backendLower || (backendLower === 'auto' && opt.value === 'auto')) {
              backendSelect.value = opt.value;
              break;
            }
          }
        }
        
        statusEl.style.background = '#064e3b';
        statusEl.style.color = '#10b981';
        statusEl.textContent = `✓ Found camera: Source ${j.index}, Backend ${j.backend}`;
        
        // Auto-connect after detection
        setTimeout(async () => {
          await connectCamera();
        }, 1000);
      } else {
        statusEl.style.background = '#7f1d1d';
        statusEl.style.color = '#ef4444';
        statusEl.textContent = '✗ Camera not found: ' + (j.error || 'No camera found. Try "List Cameras" to see available devices.');
      }
    } catch(e) {
      statusEl.style.background = '#7f1d1d';
      statusEl.style.color = '#ef4444';
      statusEl.textContent = '✗ Error: ' + (e.message || 'Failed to detect camera. Check console for details.');
      console.error('Auto-detect error:', e);
    }
  }

  // List all available cameras
  async function listAvailableCameras(){
    const listEl = document.getElementById('cameraList');
    listEl.style.display = 'block';
    listEl.style.background = '#1e293b';
    listEl.style.color = '#94a3b8';
    listEl.style.padding = '8px';
    listEl.style.borderRadius = '4px';
    listEl.textContent = '🔍 Scanning for cameras...';
    
    try {
      const r = await fetch('/list_cameras');
      const j = await r.json();
      
      if (j.ok && j.cameras && j.cameras.length > 0) {
        let html = '<strong>Available Cameras:</strong><br>';
        j.cameras.forEach(cam => {
          const typeEmoji = {
            'obs_virtual': '🎥',
            'usb_capture': '📹',
            'facetime': '📹',
            'other': '📷',
            'unknown': '❓'
          }[cam.type] || '❓';
          
          const backend = cam.backend || 'auto';
          html += `${typeEmoji} Source ${cam.index} (${backend}): ${cam.type} (${cam.size})<br>`;
        });
        listEl.innerHTML = html;
        listEl.style.background = '#064e3b';
        listEl.style.color = '#10b981';
      } else {
        listEl.style.background = '#7f1d1d';
        listEl.style.color = '#ef4444';
        listEl.textContent = '✗ No cameras found' + (j.error ? ': ' + j.error : '');
      }
    } catch(e) {
      listEl.style.background = '#7f1d1d';
      listEl.style.color = '#ef4444';
      listEl.textContent = '✗ Error: ' + (e.message || 'Failed to fetch camera list');
      console.error('List cameras error:', e);
    }
  }

  // Get system camera report
  async function getSystemCameras(){
    const listEl = document.getElementById('cameraList');
    listEl.style.display = 'block';
    listEl.style.background = '#1e293b';
    listEl.style.color = '#94a3b8';
    listEl.style.padding = '8px';
    listEl.style.borderRadius = '4px';
    listEl.textContent = '🔍 Getting system camera report...';
    
    try {
      const r = await fetch('/system_cameras');
      const j = await r.json();
      
      if (j.ok && j.system_cameras) {
        let html = '<strong>System Camera Report:</strong><br><br>';
        
        j.system_cameras.forEach(info => {
          html += `<strong>${info.method}:</strong><br>`;
          if (info.output) {
            html += `<pre style="font-size:9px;margin:4px 0;background:#0f172a;padding:4px;border-radius:2px;">${info.output}</pre>`;
          }
          if (info.error) {
            html += `<span style="color:#ef4444;">Error: ${info.error}</span><br>`;
          }
          html += '<br>';
        });
        
        listEl.innerHTML = html;
        listEl.style.background = '#064e3b';
        listEl.style.color = '#10b981';
      } else {
        listEl.style.background = '#7f1d1d';
        listEl.style.color = '#ef4444';
        listEl.textContent = '✗ Failed to get system camera report: ' + (j.error || 'Unknown error');
      }
    } catch(e) {
      listEl.style.background = '#7f1d1d';
      listEl.style.color = '#ef4444';
      listEl.textContent = '✗ Error: ' + e.message;
    }
  }

  // Reset camera system to recover from failed connections
  async function resetCameraSystem(){
    const statusEl = document.getElementById('cameraStatus');
    statusEl.style.display = 'block';
    statusEl.style.backgroundColor = '#1e293b';
    statusEl.style.color = '#e2e8f0';
    statusEl.textContent = '⏳ Resetting camera system...';
    
    try {
      const r = await fetch('/reset_camera_system', { method: 'POST' });
      const data = await r.json();
      
      if (data.ok) {
        statusEl.style.backgroundColor = '#059669';
        statusEl.style.color = '#d1fae5';
        statusEl.textContent = '✓ Camera system reset complete! Try connecting again.';
        setTimeout(() => statusEl.style.display = 'none', 5000);
      } else {
        statusEl.style.backgroundColor = '#dc2626';
        statusEl.style.color = '#fecaca';
        statusEl.textContent = 'Error: ' + (data.error || 'Reset failed');
      }
    } catch (e) {
      statusEl.style.backgroundColor = '#dc2626';
      statusEl.style.color = '#fecaca';
      statusEl.textContent = 'Error: ' + e.message;
    }
  }

  // Force restart the entire application (last resort)
  async function forceRestartApp(){
    const statusEl = document.getElementById('cameraStatus');
    statusEl.style.display = 'block';
    statusEl.style.backgroundColor = '#dc2626';
    statusEl.style.color = '#fecaca';
    statusEl.textContent = '⚠️ Force restarting Flask server in 3 seconds...';
    
    // Countdown
    for (let i = 3; i > 0; i--) {
      statusEl.textContent = `⚠️ Force restarting Flask server in ${i} seconds...`;
      await new Promise(resolve => setTimeout(resolve, 1000));
    }
    
    statusEl.textContent = '🔄 Restarting Flask server...';
    
    try {
      // Call the server restart endpoint
      const response = await fetch('/force_restart_server', { method: 'POST' });
      const data = await response.json();
      
      if (data.ok) {
        statusEl.style.backgroundColor = '#059669';
        statusEl.style.color = '#d1fae5';
        statusEl.textContent = '✓ Flask server restarting... Page will reload automatically.';
        
        // Wait a moment then reload
        setTimeout(() => {
          window.location.reload(true);
        }, 2000);
      } else {
        statusEl.textContent = 'Error: ' + (data.error || 'Restart failed');
      }
    } catch (e) {
      statusEl.textContent = 'Error: ' + e.message;
    }
  }

  // Show manual restart instructions
  function showRestartInstructions(){
    const statusEl = document.getElementById('cameraStatus');
    statusEl.style.display = 'block';
    statusEl.style.backgroundColor = '#1e293b';
    statusEl.style.color = '#e2e8f0';
    statusEl.innerHTML = `
      <strong>Manual Restart Instructions:</strong><br>
      1. Close this browser tab<br>
      2. In PowerShell/Terminal, press Ctrl+C to stop the app<br>
      3. Run: <code>python boot_cycle_gui_web-macpc-6ch.py</code><br>
      4. Open browser to: <code>http://localhost:5055</code><br>
      <em>This gives you a clean restart with fresh camera system.</em>
    `;
    statusEl.style.fontSize = '10px';
    statusEl.style.lineHeight = '1.4';
  }

  async function probe(){
    const body = {
      src: document.getElementById('src').value,
      backend: document.getElementById('backend').value,
      fourcc: document.getElementById('fourcc').value,
      res: document.getElementById('res').value
    };
    const r = await fetch('/probe', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(body)
    });
    document.getElementById('probeOut').textContent = JSON.stringify(await r.json(), null, 2);
  }

  // Equipment modal functions
  function showEquipModal(){
    document.getElementById('equipModal').classList.add('show');
  }
  
  function hideEquipModal(){
    document.getElementById('equipModal').classList.remove('show');
  }
  
  async function saveEquipmentAndContinue(){
    const data = {
      console1_serial: document.getElementById('console1_serial').value,
      console1_scope: document.getElementById('console1_scope').value,
      console2_serial: document.getElementById('console2_serial').value,
      console2_scope: document.getElementById('console2_scope').value,
      console3_serial: document.getElementById('console3_serial').value,
      console3_scope: document.getElementById('console3_scope').value
    };
    
    try {
      await fetch('/set_equipment', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify(data)
      });
      hideEquipModal();
      updateEquipmentDisplay();
    } catch(e) {
      console.error('Error saving equipment:', e);
    }
  }
  
  async function updateEquipmentDisplay(){
    try {
      const r = await fetch('/get_equipment');
      const j = await r.json();
      const eq = j.equipment;
      
      let summary = '';
      summary += `C1: ${eq.console1.serial || '?'} / ${eq.console1.scope_id || '?'}<br>`;
      summary += `C2: ${eq.console2.serial || '?'} / ${eq.console2.scope_id || '?'}<br>`;
      summary += `C3: ${eq.console3.serial || '?'} / ${eq.console3.scope_id || '?'}`;
      
      document.getElementById('equipSummary').innerHTML = summary;
      document.getElementById('equipInfo').style.display = 'block';
      
      // Update modal fields
      document.getElementById('console1_serial').value = eq.console1.serial || '';
      document.getElementById('console1_scope').value = eq.console1.scope_id || '';
      document.getElementById('console2_serial').value = eq.console2.serial || '';
      document.getElementById('console2_scope').value = eq.console2.scope_id || '';
      document.getElementById('console3_serial').value = eq.console3.serial || '';
      document.getElementById('console3_scope').value = eq.console3.scope_id || '';
    } catch(e) {
      console.error('Error updating equipment display:', e);
    }
  }

  function setPill(el, det, idx, cnt, discnt){
    const map = {
      INTERFACE: ['ok', 'Scope Connected'],
      BARS: ['bad', 'Scope Disconnected'],
      NO_SIGNAL: ['nosig', 'No Signal'],
      OTHER: ['other', 'Other']
    };
    const [cls, label] = map[det] || ['other', 'Other'];
    el.className = 'pill ' + cls;
    const cntVal = (typeof cnt === 'number') ? cnt : 0;
    const discntVal = (typeof discnt === 'number') ? discnt : 0;
    el.querySelector('span').textContent = label + ': ' + cntVal + '/' + discntVal;
  }

  async function refreshPills(){
    try {
      const r = await fetch('/grid_status');
      const j = await r.json();
      const arr = j.tiles || [];
      for (let i = 0; i < 6; i++) {
        const videoNum = TILE_TO_PILL[i];            // 1,3,5,2,4,6
        const el = document.getElementById('p' + videoNum);
        if (!el) continue;
        const det = (arr[i] && arr[i].det) || 'OTHER';
        const cnt = (arr[i] && typeof arr[i].cnt === 'number') ? arr[i].cnt : 0;
        const discnt = (arr[i] && typeof arr[i].discnt === 'number') ? arr[i].discnt : 0;
        setPill(el, det, videoNum, cnt, discnt);
      }
    } catch (e) {
      console.error(e);
    }
  }

  // Inserted resetTallies after probe
  async function resetTallies(){
    try{
      await fetch('/reset_tallies', { method: 'POST' });
      await refreshPills();
    }catch(e){
      console.error(e);
    }
  }

  setInterval(() => {
    const im = document.getElementById('grid');
    if (im) im.src = bust('/thumb');
  }, 500);

    setInterval(refreshPills, 800);
    refreshPills();
    
    async function refreshCSVInfo(){
  try{
    const r = await fetch('/status'); const j = await r.json();
    const name = (j.csv || '').split(/[\\/]/).pop() || '-';
    const el = document.getElementById('csvName');
    if (el) el.textContent = name;
    const dl = document.getElementById('dl');
    if (dl) dl.href = '/download';
    
    // Update test indicator based on detection_active status
    const indicator = document.getElementById('testIndicator');
    if (indicator) {
      if (j.detection_active) {
        indicator.classList.add('running');
        // Force visibility with important
        indicator.style.setProperty('opacity', '1', 'important');
        indicator.style.setProperty('display', 'inline-block', 'important');
        indicator.style.setProperty('background', '#ef4444', 'important');
      } else {
        indicator.classList.remove('running');
        indicator.style.setProperty('opacity', '0', 'important');
      }
    }
  }catch(e){}
}
setInterval(refreshCSVInfo, 1200);
refreshCSVInfo();

// Initialize equipment display on load
updateEquipmentDisplay();

// Auto-start video feed when page loads
startVideo();
</script></body></html>"""


# Live adjust endpoint
@app.route("/adjust", methods=["POST"])
def adjust():
    cfg = request.get_json(silent=True) or {}
    with mon.lock:
        if "white_pct" in cfg:
            mon.white_frac_gate = cfg.get("white_pct", mon.white_frac_gate)
        if "mean_gate" in cfg:
            mon.mean_gate = cfg.get("mean_gate", mon.mean_gate)
        if "roi_inset" in cfg:
            try:
                mon.roi_inset_px = int(cfg.get("roi_inset", getattr(mon, "roi_inset_px", 0)) or 0)
            except Exception:
                pass
        current = {
            "white_frac_gate": mon.white_frac_gate,
            "mean_gate": mon.mean_gate,
            "roi_inset": getattr(mon, "roi_inset_px", 0)
        }
    return jsonify(ok=True, **current)

@app.post("/set_equipment")
def set_equipment():
    """Set console serial numbers and scope IDs."""
    cfg = request.get_json(silent=True) or {}
    with mon.lock:
        # Update equipment info
        for i in range(1, 4):
            console_key = f"console{i}"
            serial = cfg.get(f"console{i}_serial", "").strip()
            scope_id = cfg.get(f"console{i}_scope", "").strip()
            if console_key in mon.equipment:
                mon.equipment[console_key]["serial"] = serial
                mon.equipment[console_key]["scope_id"] = scope_id
    return jsonify(ok=True, equipment=mon.equipment)

@app.get("/get_equipment")
def get_equipment():
    """Get current equipment configuration."""
    with mon.lock:
        return jsonify(equipment=mon.equipment)

@app.get("/ping")
def ping():
    return Response("pong", mimetype="text/plain")



@app.get("/status")
def status():
    with mon.lock:
        sf = label_for(mon.status) if mon.status in ("BARS","INTERFACE","OTHER","NO_SIGNAL") else mon.status
        return jsonify(
            status=mon.status, status_friendly=sf, running=mon.running,
            video_running=mon.video_running, detection_active=mon.detection_active,
            count_bars=mon.count_bars, count_int=mon.count_int,
            count_other=mon.count_other, cycles=mon.cycles,
            last_seen=mon.last_seen, backend=mon.backend_name,
            fourcc=mon.fourcc, res_preset=mon.res_preset,
            csv=mon.csv_path, white_frac_gate=mon.white_frac_gate,
            mean_gate=mon.mean_gate,
            roi_inset=getattr(mon, "roi_inset_px", 0)
        )


# Helper: draw annotated 3x2 grid with per-tile status/labels
def _annotated_grid_from_frame(bgr):
    """
    Draw a 3x2 annotated grid on the latest frame:
      - For each tile, run the same ROI-based decision as main loop
      - Paint a status badge (Connected / Not Connected / No Signal / Other)
      - Overlay the tile label: VIDEO 1..6
    This is purely for UI; it does not mutate counters.
    """
    # Ensure we have reference hashes; compute lazily if needed
    with mon.lock:
        bars_h_roi = mon._bars_h_roi
        int_h_list = mon._int_h_list
        ref_mean   = mon._ref_mean
        ref_bright = mon._ref_bright
        cw         = mon.center_w
        thr        = mon.thresh
        dmean      = mon.dark_mean
        dstd       = mon.dark_std
        margin     = mon.margin

    if bars_h_roi is None or not int_h_list:
        try:
            paths = discover_int_refs()
            bars_h_roi = ph_bars_ref_roi(mon.bars_ref)
            int_h_list = ph_int_ref_list(paths) or [ph_int_ref(mon.int_ref)]
            m, s, b = ref_roi_stats_from_paths(paths)
            ref_mean, ref_bright = m, b
            with mon.lock:
                mon._bars_h_roi = bars_h_roi
                mon._int_h_list = int_h_list
                mon._ref_mean = ref_mean
                mon._ref_bright = ref_bright
        except Exception:
            pass

    tiles, rects = split_grid(bgr, GRID_COLS, GRID_ROWS)
    out = bgr.copy()

    # Map row-major tile index -> displayed VIDEO number (1,3,5 on top; 2,4,6 on bottom)
    TILE_TO_VIDEO = [1, 3, 5, 2, 4, 6]

    # Colors (B,G,R)
    C_OK   = ( 22, 163,  74)   # green
    C_BAD  = ( 38,  38, 255)   # red
    C_WARN = ( 18, 182, 252)   # cyan-ish for "Other"
    C_NO   = ( 64,  64,  64)   # gray

    tiles_out = []
    for idx, (tile, (x, y, w, h)) in enumerate(zip(tiles, rects), start=1):
        det, db, di, _cg, mean_l, std_l, roi_mean, roi_bright, bright_ok, white_frac_gray, white_frac_rgb = decide(
            tile, bars_h_roi, int_h_list, cw, thr, dmean, dstd, margin, ref_mean, ref_bright,
            white_frac_override=getattr(mon, "white_frac_gate", None),
            mean_gate_override=getattr(mon, "mean_gate", None)
        )
        try:
            discnt_val = int(mon.tile_disconnected_counts[idx-1])
        except Exception:
            discnt_val = 0
        tiles_out.append({"det": det, "db": int(db), "di": int(di), "cnt": int(mon.tile_counts[idx-1]), "discnt": discnt_val})
        # --- Per-tile transition tracking and CSV logging ---
        try:
            tile_idx0 = idx - 1
            prev = mon.tile_last[tile_idx0]
            TILE_TO_VIDEO = [1, 3, 5, 2, 4, 6]
            video_num = TILE_TO_VIDEO[tile_idx0]
            
            # Track state changes
            if det != prev:
                # Record state history
                mon.tile_state_history[tile_idx0].append({
                    "state": det,
                    "timestamp": time.time()
                })
                
                # Calculate elapsed time from test start
                elapsed_from_test_start = None
                if mon.test_start_time:
                    elapsed_from_test_start = (time.time() - mon.test_start_time.timestamp())
                
                # Track timing for Disconnected → Connected transitions
                if det == "BARS":
                    # Mark start of Scope Disconnected period
                    # Don't increment count yet - wait to see if next state is NO_SIGNAL
                    mon.tile_disconnected_start[tile_idx0] = time.time()
                    
                    # Check if we entered BARS from NO_SIGNAL (potential partial cycle)
                    if prev == "NO_SIGNAL":
                        mon.tile_in_partial_cycle[tile_idx0] = True
                    
                    # Log Scope Disconnected with elapsed time and current cycle number
                    _csv_append(video_num, det, 
                               elapsed_secs=elapsed_from_test_start,
                               cycle_num=mon.tile_complete_cycles[tile_idx0],
                               partial_cycle_num=None,
                               bars_dist=int(db), int_dist=int(di))
                    
                elif det == "INTERFACE":
                    # Scope Connected - increment cycle number FIRST
                    mon.tile_complete_cycles[tile_idx0] += 1
                    
                    # If we were in a partial cycle but went to INTERFACE instead of NO_SIGNAL, cancel the partial cycle
                    if mon.tile_in_partial_cycle[tile_idx0]:
                        mon.tile_in_partial_cycle[tile_idx0] = False
                    
                    # Calculate elapsed time if coming from Disconnected (for reconnection timing)
                    reconnection_elapsed = None
                    if mon.tile_disconnected_start[tile_idx0] is not None:
                        reconnection_elapsed = time.time() - mon.tile_disconnected_start[tile_idx0]
                        mon.tile_cycle_times[tile_idx0].append(reconnection_elapsed)
                        mon.tile_disconnected_start[tile_idx0] = None
                    
                    mon.tile_counts[tile_idx0] += 1
                    
                    # Log Scope Connected with elapsed time from test start and cycle number
                    _csv_append(video_num, det, 
                               elapsed_secs=elapsed_from_test_start,
                               cycle_num=mon.tile_complete_cycles[tile_idx0],
                               partial_cycle_num=None,
                               bars_dist=int(db), int_dist=int(di))
                    
                elif det == "NO_SIGNAL":
                    # Check for partial cycle completion: NO_SIGNAL → BARS → NO_SIGNAL
                    partial_cycle_count = None
                    if prev == "BARS" and mon.tile_in_partial_cycle[tile_idx0]:
                        # Complete partial cycle detected: NO_SIGNAL → BARS → NO_SIGNAL
                        mon.tile_partial_cycles[tile_idx0] += 1
                        partial_cycle_count = mon.tile_partial_cycles[tile_idx0]
                        mon.tile_in_partial_cycle[tile_idx0] = False
                    
                    # Only increment disconnected count if we're transitioning FROM BARS to NO_SIGNAL
                    # This means the scope was disconnected and went to no signal (complete disconnect cycle)
                    # If it went BARS → INTERFACE, that's a normal cycle, not a disconnect
                    if prev == "BARS":
                        mon.tile_disconnected_counts[tile_idx0] += 1
                        # Clear the disconnected start time since we've completed the disconnect cycle
                        if mon.tile_disconnected_start[tile_idx0] is not None:
                            mon.tile_disconnected_start[tile_idx0] = None
                    
                    # Log NO_SIGNAL - no elapsed_secs, but include partial_cycle_count if this completes a partial cycle
                    _csv_append(video_num, det, 
                               elapsed_secs=None,  # Don't log elapsed_secs for NO_SIGNAL
                               cycle_num=None,
                               partial_cycle_num=partial_cycle_count,
                               bars_dist=int(db), int_dist=int(di))
                    
                else:  # OTHER
                    # If we were in a partial cycle but went to OTHER, cancel the partial cycle
                    if mon.tile_in_partial_cycle[tile_idx0]:
                        mon.tile_in_partial_cycle[tile_idx0] = False
                    
                    # Don't log any values (elapsed_secs, cycle_num, partial_cycle_num) for OTHER state
                    _csv_append(video_num, det, 
                               elapsed_secs=None,  # No elapsed_secs for OTHER
                               cycle_num=None,  # No cycle_num for OTHER
                               partial_cycle_num=None,  # No partial_cycle_num for OTHER
                               bars_dist=int(db), int_dist=int(di))
                
                mon.tile_last[tile_idx0] = det
        except Exception as e:
            print(f"[Tile tracking] Error: {e}")
            pass
        if det == "INTERFACE":
            color, text = C_OK, "Scope Connected"
        elif det == "BARS":
            color, text = C_BAD, "Scope Disconnected"
        elif det == "NO_SIGNAL":
            color, text = C_NO, "No Signal"
        else:
            color, text = C_WARN, "Other"

        # draw status badge inside tile
        _ = draw_status_badge(out, x + 10, y + 10, text, color)

        # --- Debug overlay: show parameters and measured ROI values ---
        try:
            ov_frac = getattr(mon, "white_frac_gate", None)
            ov_mean = getattr(mon, "mean_gate", None)
            debug_lines = [
                f"roi_inset={getattr(mon,'roi_inset_px',0)}",
                f"ov_frac={ov_frac}",
                f"ov_mean={ov_mean}",
                f"roi_mean={roi_mean:.1f}",
                f"w_frac_g={white_frac_gray:.3f}",
                f"w_frac_rgb={white_frac_rgb:.3f}",
                f"bright_ok={bright_ok}"
            ]
            tx = x + 10
            ty = y + 40
            font = cv2.FONT_HERSHEY_SIMPLEX
            scale = 0.5
            thickness = 1
            for i, line in enumerate(debug_lines):
                y0 = ty + i * 12
                cv2.putText(out, line, (tx, y0), font, scale, (200, 200, 200), thickness, cv2.LINE_AA)
        except Exception:
            pass


        # overlay the tile label centered near top-right of the tile
        label = f"VIDEO {TILE_TO_VIDEO[idx-1]}"
        try:
            font = cv2.FONT_HERSHEY_SIMPLEX
            scale = 1.0
            thickness = 2
            (tw, th), _ = cv2.getTextSize(label, font, scale, thickness)
            cx = x + w - tw - 14
            cy = y + 18 + th
            cv2.putText(out, label, (cx, cy), font, scale, (255, 255, 255), thickness, cv2.LINE_AA)
        except Exception:
            pass

        # draw the detected ROI box for visual confirmation
        try:
            # map ROI (computed in tile space) back to out image coordinates
            tx, ty, rw, rh = _roi_box_for_frame(w, h, getattr(mon, "roi_inset_px", 0))
            cv2.rectangle(out, (x + tx, y + ty), (x + tx + rw, y + ty + rh), (255, 255, 255), 1)
        except Exception:
            pass

    # Cache the detections snapshot for UI consistency (used by /grid_status)
    try:
        with mon.lock:
            mon.last_grid = {"ts": time.time(), "tiles": tiles_out}
    except Exception:
        pass

    return out

@app.get("/thumb")
def thumb():
    with mon.lock:
        frame = mon.last_frame

    if frame is not None:
        try:
            # Annotate the live 3x2 grid with per‑tile status/labels
            annotated = _annotated_grid_from_frame(frame)
            img = annotated
        except Exception:
            # fall back to simple composite if annotation fails
            img = _composite_thumb_from_frame(frame)
    else:
        img = _placeholder_thumb(1200, 675)

    ok, buf = cv2.imencode(".jpg", img)
    if not ok:
        return Response(status=500)
    return Response(buf.tobytes(), mimetype="image/jpeg")

# New endpoint: grid_status for per-tile detections (JSON)
@app.get("/grid_status")
def grid_status():
    # First, serve the cached grid computed during /thumb rendering (same frame)
    with mon.lock:
        cached = mon.last_grid
        frame = mon.last_frame
        bars_h_roi = mon._bars_h_roi
        int_h_list = mon._int_h_list
        ref_mean   = mon._ref_mean
        ref_bright = mon._ref_bright
        cw         = mon.center_w
        thr        = mon.thresh
        dmean      = mon.dark_mean
        dstd       = mon.dark_std
        margin     = mon.margin

    if cached and isinstance(cached, dict) and cached.get("tiles"):
        return jsonify(cached)

    # If no cache yet, compute once from the current frame and cache it
    if frame is None:
        return jsonify(tiles=[])

    # Ensure refs
    if bars_h_roi is None or not int_h_list:
        try:
            paths = discover_int_refs()
            bars_h_roi = ph_bars_ref_roi(mon.bars_ref)
            int_h_list = ph_int_ref_list(paths) or [ph_int_ref(mon.int_ref)]
            m, s, b = ref_roi_stats_from_paths(paths)
            ref_mean, ref_bright = m, b
            with mon.lock:
                mon._bars_h_roi = bars_h_roi
                mon._int_h_list = int_h_list
                mon._ref_mean = ref_mean
                mon._ref_bright = ref_bright
        except Exception:
            pass

    tiles, rects = split_grid(frame, GRID_COLS, GRID_ROWS)
    tiles_out = []
    for i, tile in enumerate(tiles):
        det, db, di, _cg, mean_l, std_l, roi_mean, roi_bright, bright_ok, _wfg, _wfr = decide(
            tile, bars_h_roi, int_h_list, cw, thr, dmean, dstd, margin, ref_mean, ref_bright,
            white_frac_override=getattr(mon, "white_frac_gate", None),
            mean_gate_override=getattr(mon, "mean_gate", None)
        )
        try:
            cnt_val = int(mon.tile_counts[i])
        except Exception:
            cnt_val = 0
        try:
            discnt_val = int(mon.tile_disconnected_counts[i])
        except Exception:
            discnt_val = 0
        tiles_out.append({"det": det, "db": int(db), "di": int(di), "cnt": cnt_val, "discnt": discnt_val})

    snapshot = {"ts": time.time(), "tiles": tiles_out}
    try:
        with mon.lock:
            mon.last_grid = snapshot
    except Exception:
        pass
    return jsonify(snapshot)

@app.get("/peek")
def peek():
    """Return quick diagnostic metrics or perform a one-shot detection."""
    # If loop is already running and metrics exist, just return them
    with mon.lock:
        m = mon.last_metrics.copy()
        have_recent = mon.last_frame is not None
        src = mon.src
        stream = mon.stream_path
        cw = mon.center_w
        thr = mon.thresh
        dmean = mon.dark_mean
        dstd = mon.dark_std
        bars_path = mon.bars_ref
        int_path = mon.int_ref
        backend_name = mon.backend_name
        fourcc_sel = (mon.fourcc or "auto").upper()
        res_preset = (mon.res_preset or "1080p").lower()

    if have_recent:
        for k in ("db", "di", "mean", "std"):
            if m.get(k) is None:
                m[k] = "-"
        return jsonify(m)

    # Load reference hashes and paths
    try:
        int_ref_paths = discover_int_refs()
        bars_h_roi = ph_bars_ref_roi(bars_path)
        int_h_list = ph_int_ref_list(int_ref_paths) or [ph_int_ref(int_path)]
        ref_mean, ref_std, ref_bright = ref_roi_stats_from_paths(int_ref_paths)
    except Exception as e:
        return jsonify(error=f"ref load: {e}")

    use_stream = isinstance(stream, str) and stream.strip() != ""

    def _open_with_backend(_src, _backend):
        # URL/file path
        if isinstance(_src, str) and not _src.isdigit():
            return cv2.VideoCapture(_src), _src
        # numeric index
        try:
            idx = int(_src)
        except Exception:
            idx = _src
        return cv2.VideoCapture(idx, _backend), idx

    # Resolve backend const
    if (backend_name or "auto").lower() == "auto":
        backend = (
            cv2.CAP_MSMF if platform.system() == "Windows"
            else cv2.CAP_AVFOUNDATION if platform.system() == "Darwin"
            else cv2.CAP_V4L2
        )
    else:
        name = backend_name.lower()
        backend = (
            cv2.CAP_MSMF if name == "msmf" else
            cv2.CAP_DSHOW if name == "dshow" else
            cv2.CAP_AVFOUNDATION if name == "avfoundation" else
            cv2.CAP_V4L2
        )

    # Desired resolution
    want_w, want_h = (1280, 720) if res_preset == "720p" else (1920, 1080)

    cap = None
    used_idx = None
    used_backend = backend
    try:
        if use_stream:
            cap = cv2.VideoCapture(stream.strip())
            used_idx = stream.strip()
        else:
            # first attempt
            cap, used_idx = _open_with_backend(src, backend)

            # fallbacks
            if (not cap) or (not cap.isOpened()):
                if platform.system() == "Windows":
                    alt_backend = cv2.CAP_DSHOW if backend == cv2.CAP_MSMF else cv2.CAP_MSMF
                    if cap: cap.release()
                    cap, used_idx = _open_with_backend(src, alt_backend)
                    used_backend = alt_backend
                elif platform.system() == "Darwin":
                    if cap: cap.release()
                    fourcc_sel = (mon.fourcc or "auto").upper()
                    want_w, want_h = (1280, 720) if (mon.res_preset or "1080p").lower()=="720p" else (1920, 1080)
                    cap, used_idx = _mac_try_indices_for_nonblack([1,0,2,3,4,5], cv2.CAP_AVFOUNDATION, want_w, want_h, fourcc_sel)
                    used_backend = cv2.CAP_AVFOUNDATION
                    if cap is None:
                        # fallback to user‑selected index even if black
                        try:
                            cap = cv2.VideoCapture(int(src), cv2.CAP_AVFOUNDATION)
                            used_idx = int(src)
                        except Exception:
                            cap = None
                            used_idx = None

        if not cap or not cap.isOpened():
            with mon.lock:
                mon.status = "error: cannot open source"
                mon.running = False
            return jsonify(error="cannot open source")

        # Normalize + apply preferences
        try: cap.set(cv2.CAP_PROP_CONVERT_RGB, 1)
        except Exception: pass

        if fourcc_sel == "MJPG":
            try: cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'MJPG'))
            except Exception: pass
        elif fourcc_sel == "YUY2":
            try: cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'YUY2'))
            except Exception: pass

        try:
            cap.set(cv2.CAP_PROP_FRAME_WIDTH,  want_w)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, want_h)
        except Exception:
            pass
        try: cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        except Exception: pass
        try: cap.set(cv2.CAP_PROP_FPS, 30)
        except Exception: pass

        # Warm up a few frames
        for _ in range(5):
            cap.read()

        # Try reading a frame
        ok, frame = cap.read()
        if not ok or frame is None:
            return jsonify(error="cannot read frame")
        # Flag obviously black frames to aid debugging in the UI
        try:
            gray_chk = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            mean_chk = float(np.mean(gray_chk))
            black_hint = (mean_chk < 5.0)
        except Exception:
            black_hint = False

        # --- Compute per-reference distances for debugging ---
        roi = roi_connected_gray(frame)
        roi_eq = _equalize_hist(roi)
        phv_int = ih.phash(Image.fromarray(roi_eq))
        distances = []
        for h in int_h_list:
            try:
                distances.append(int(phv_int - h))
            except Exception:
                distances.append(None)
        di = min([d for d in distances if d is not None] or [999])

        # Decide state (ROI-only logic)
        det, db, _, cg, mean_l, std_l, roi_mean, roi_bright, bright_ok, _wfg, _wfr = decide(
            frame, bars_h_roi, int_h_list, cw, thr, dmean, dstd, mon.margin, ref_mean, ref_bright,
            white_frac_override=getattr(mon, "white_frac_gate", None),
            mean_gate_override=getattr(mon, "mean_gate", None)
        )

        # Size/FPS/FourCC readback (best-effort)
        try:
            h, w = frame.shape[:2]
            size_str = f"{w}x{h}"
        except Exception:
            size_str = "-"
        try:
            v = cap.get(cv2.CAP_PROP_FOURCC)
            fourcc_read = _fourcc_to_str(int(v)) if v and v != 0 else ""
        except Exception:
            fourcc_read = ""
        try:
            fps_g = cap.get(cv2.CAP_PROP_FPS)
            fps_val = round(float(fps_g), 1) if fps_g and fps_g > 0 else None
        except Exception:
            fps_val = None

        return jsonify(
            det=det,
            db=int(db), di=int(di),
            mean=round(mean_l, 2), std=round(std_l, 2),
            distances=distances,
            roi_mean=round(float(roi_mean), 1),
            roi_bright=round(float(roi_bright), 3),
            bright_ok=bool(bright_ok),
            backend_used=int(used_backend) if isinstance(used_backend, int) else used_backend,
            index_used=used_idx,
            size=size_str,
            fourcc=fourcc_read or None,
            fps=fps_val,
            black_hint=black_hint
        )

    finally:
        try:
            if cap: cap.release()
        except Exception:
            pass
# Helper: Auto-detect OBS Virtual Camera or first available camera
def identify_camera_type(cap, idx):
    """
    Try to identify camera type based on properties.
    Returns: "obs_virtual", "facetime", "usb_capture", "other", or "unknown"
    """
    try:
        # Get camera properties
        width = cap.get(cv2.CAP_PROP_FRAME_WIDTH)
        height = cap.get(cv2.CAP_PROP_FRAME_HEIGHT)
        fps = cap.get(cv2.CAP_PROP_FPS)
        
        # Try to get backend name
        try:
            backend = cap.getBackendName()
        except:
            backend = "unknown"
        
        # More conservative identification - don't guess OBS unless we're sure
        is_standard_resolution = (width, height) in [(1920, 1080), (1280, 720)]
        is_small_resolution = (width, height) in [(640, 480), (320, 240), (160, 120)]
        is_standard_fps = fps >= 25 and fps <= 60
        
        if platform.system() == "Darwin":  # macOS
            # Check if OBS is running - if so, higher indices are more likely to be OBS Virtual Camera
            obs_running = False
            try:
                import subprocess
                result = subprocess.run(['pgrep', '-f', 'obs'], capture_output=True, timeout=2)
                obs_running = result.returncode == 0
            except:
                pass
            
            if idx == 0:
                return "facetime"  # FaceTime camera is almost always at index 0
            elif obs_running and idx >= 1:
                # If OBS is running, indices 1+ are likely OBS Virtual Camera
                # OBS Virtual Camera typically appears at index 1 or 2
                if is_standard_resolution and is_standard_fps:
                    return "obs_virtual"  # Likely OBS Virtual Camera
            elif is_small_resolution:
                return "other"  # Small cameras
            elif is_standard_resolution and is_standard_fps:
                return "usb_capture"  # Likely USB capture device
            else:
                return "other"
        else:  # Windows/Linux
            # Windows: Check if OBS is running to identify OBS Virtual Camera
            if platform.system() == "Windows":
                obs_running = False
                try:
                    import subprocess
                    result = subprocess.run(['tasklist', '/FI', 'IMAGENAME eq obs64.exe'], 
                                          capture_output=True, timeout=2)
                    obs_running = result.returncode == 0 and b'obs64.exe' in result.stdout
                except:
                    pass
                
                # If OBS is running and camera has standard resolution, prioritize it as OBS Virtual Camera
                # OBS Virtual Camera appears as a high-resolution camera (720p/1080p)
                if obs_running and is_standard_resolution:
                    # OBS Virtual Camera typically appears at index 1 or 2, but can be anywhere
                    # Prefer indices 1+ over 0 for OBS (index 0 is usually built-in camera)
                    if idx >= 1:
                        # High-res camera at index 1+ when OBS is running = likely OBS Virtual Camera
                        return "obs_virtual"  # Likely OBS Virtual Camera
                    elif idx == 0:
                        # Index 0 could be OBS if it's high-res (but usually built-in is 640x480)
                        # Only mark as OBS if it's truly high-res (not small resolution)
                        if not is_small_resolution:
                            return "obs_virtual"  # Could be OBS Virtual Camera
                
                # Windows: Better detection of USB capture devices (only if not OBS)
                if idx == 0 and is_small_resolution:
                    return "other"  # Built-in laptop camera (usually 640x480)
                elif is_small_resolution:
                    return "other"  # Small cameras
                elif is_standard_resolution:
                    # Standard resolution (720p/1080p) - could be USB capture if OBS not running
                    # OR could be OBS if OBS is running but we didn't catch it above
                    if obs_running:
                        # If OBS is running and we have standard res, prefer OBS over USB
                        # (This handles edge cases where OBS check above didn't match)
                        if idx >= 1:
                            return "obs_virtual"  # Prefer OBS when OBS is running
                        else:
                            return "usb_capture"  # Could be USB at index 0
                    else:
                        # OBS not running - standard res is likely USB capture
                        return "usb_capture"
                else:
                    return "other"
            else:
                # Not Windows - use generic logic
                if idx == 0 and is_small_resolution:
                    return "other"
                elif is_small_resolution:
                    return "other"
                elif is_standard_resolution:
                    return "usb_capture"
                else:
                    return "other"
                
    except Exception:
        return "unknown"

def reset_camera_system():
    """Reset the camera system to recover from failed connections."""
    print("[reset_camera_system] Resetting camera system...")
    
    # Stop any running video threads first
    with mon.lock:
        if mon.video_running:
            print("[reset_camera_system] Stopping video thread...")
            mon.video_running = False
            mon.detection_active = False
            mon.running = False
            mon.status = "camera_resetting"
    
    # Wait for thread to stop
    import time
    time.sleep(2.0)
    
    # Force release any stuck camera handles
    try:
        # Try to release any global camera objects
        if hasattr(mon, 'cap') and mon.cap:
            mon.cap.release()
            mon.cap = None
    except:
        pass
    
    # Force garbage collection multiple times
    import gc
    for _ in range(3):
        gc.collect()
        time.sleep(0.5)
    
    # Additional Windows-specific cleanup
    if platform.system() == "Windows":
        try:
            import subprocess
            # Kill any stuck camera processes (non-destructive)
            subprocess.run(["taskkill", "/F", "/IM", "BootCycleLogger.exe"], 
                         capture_output=True, timeout=5)
        except:
            pass
    
    print("[reset_camera_system] ✓ Camera system reset complete")

def auto_detect_camera():
    """
    Auto-detect the best camera source for the system.
    Prioritizes OBS Virtual Camera detection across all platforms.
    Returns: (source_index, backend_code, backend_name) or (0, CAP_ANY, "auto")
    """
    print("[Camera Detection] Starting auto-detection...")
    
    if platform.system() == "Windows":
        # Windows: Check if OBS is running first
        obs_running = False
        try:
            import subprocess
            result = subprocess.run(['tasklist', '/FI', 'IMAGENAME eq obs64.exe'], 
                                  capture_output=True, timeout=2)
            obs_running = result.returncode == 0 and b'obs64.exe' in result.stdout
            if obs_running:
                print("[Camera Detection] ✓ OBS is running - will prioritize OBS Virtual Camera")
            else:
                print("[Camera Detection] ⚠ OBS is not running")
        except Exception as e:
            print(f"[Camera Detection] Could not check OBS status: {e}")
        
        # Windows: Try OBS Virtual Camera by name first (only if OBS is running)
        if obs_running:
            print("[Camera Detection] Checking for OBS Virtual Camera by name...")
            
            # Try OBS Virtual Camera by name with multiple backends
            obs_names = [
                "OBS Virtual Camera", 
                "OBS-Camera", 
                "OBS Virtual Source",
                "OBS Virtual Camera (DirectShow)",
                "OBS Virtual Camera (Media Foundation)"
            ]
            
            # Try DSHOW first on Windows (more reliable for OBS)
            obs_backends = [
                ("DSHOW", cv2.CAP_DSHOW),
                ("MSMF", cv2.CAP_MSMF),
                ("ANY", cv2.CAP_ANY)
            ]
            
            for obs_name in obs_names:
                for backend_name, backend_code in obs_backends:
                    cap = None
                    try:
                        print(f"[Camera Detection] Trying OBS Virtual Camera: '{obs_name}' with {backend_name}...")
                        cap = cv2.VideoCapture(obs_name, backend_code)
                        if cap.isOpened():
                            # Set buffer to 1 for faster response
                            try:
                                cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
                            except:
                                pass
                            
                            # Try to read frame with timeout (prevent hanging on problematic cameras)
                            frame_result = [None, None]
                            frame_ready = threading.Event()
                            
                            def read_frame():
                                try:
                                    ret, frame = cap.read()
                                    frame_result[0] = ret
                                    frame_result[1] = frame
                                except Exception as e:
                                    frame_result[0] = False
                                    frame_result[1] = None
                                finally:
                                    frame_ready.set()
                            
                            read_thread = threading.Thread(target=read_frame, daemon=True)
                            read_thread.start()
                            frame_ready.wait(timeout=2.0)  # 2 second timeout
                            
                            if frame_ready.is_set():
                                ret, frame = frame_result[0], frame_result[1]
                                if ret and frame is not None:
                                    h, w = frame.shape[:2]
                                    if cap:
                                        cap.release()
                                    print(f"[Camera Detection] ✓ Found OBS Virtual Camera: '{obs_name}', backend={backend_name}, size={w}x{h}")
                                    return obs_name, backend_code, backend_name
                            
                            if cap:
                                cap.release()
                    except Exception as e:
                        if cap:
                            try:
                                cap.release()
                            except:
                                pass
                        print(f"[Camera Detection] OBS Virtual Camera '{obs_name}' with {backend_name} failed: {e}")
                        continue
        
        # Windows: Scan indices and identify camera types
        # Prioritize OBS Virtual Camera, then USB capture devices
        backends = [
            ("DSHOW", cv2.CAP_DSHOW),
            ("MSMF", cv2.CAP_MSMF)
        ]
        
        found_cameras = []
        
        # Scan indices 0-10 first (most systems have cameras in this range)
        # This prevents hanging on high indices that don't exist
        max_initial_scan = 10
        
        excellent_camera_found = False
        for backend_name, backend_code in backends:
            # Skip this backend if we already found excellent cameras (OBS or high-res USB) with previous backend
            if excellent_camera_found:
                print(f"[Camera Detection] Already found excellent camera, skipping {backend_name} backend")
                break
            
            for idx in range(max_initial_scan + 1):  # 0-10
                try:
                    print(f"[Camera Detection] Testing index {idx} with {backend_name}...")
                    cap = None
                    try:
                        cap = cv2.VideoCapture(idx, backend_code)
                        
                        if cap.isOpened():
                            # Set buffer to 1 for faster response
                            try:
                                cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
                            except:
                                pass
                            
                            # Try to read a frame with timeout on Windows (this can hang)
                            # Use threading to implement timeout
                            frame_result = [None, None]
                            frame_ready = threading.Event()
                            
                            def read_frame():
                                try:
                                    ret, frame = cap.read()
                                    frame_result[0] = ret
                                    frame_result[1] = frame
                                except Exception as e:
                                    frame_result[0] = False
                                    frame_result[1] = None
                                finally:
                                    frame_ready.set()
                            
                            read_thread = threading.Thread(target=read_frame, daemon=True)
                            read_thread.start()
                            frame_ready.wait(timeout=2.0)  # 2 second timeout
                            
                            if not frame_ready.is_set():
                                print(f"[Camera Detection] Timeout reading from index {idx} with {backend_name}")
                                if cap:
                                    cap.release()
                                continue
                            
                            ret, frame = frame_result[0], frame_result[1]
                            if ret and frame is not None:
                                h, w = frame.shape[:2]
                                # Try to identify camera type
                                camera_type = identify_camera_type(cap, idx)
                                found_cameras.append({
                                    'index': idx,
                                    'backend': backend_code,
                                    'backend_name': backend_name,
                                    'type': camera_type,
                                    'size': f"{w}x{h}",
                                    'width': w,
                                    'height': h
                                })
                                print(f"[Camera Detection] Found camera {idx} ({backend_name}): {camera_type} ({w}x{h})")
                                
                                # If we found an excellent camera (OBS or high-res USB), stop scanning
                                if camera_type in ['obs_virtual', 'usb_capture'] and w >= 1280:
                                    print(f"[Camera Detection] Found excellent camera, stopping scan early")
                                    excellent_camera_found = True
                                    if cap:
                                        cap.release()
                                    break
                        if cap:
                            cap.release()
                    except Exception as inner_e:
                        if cap:
                            try:
                                cap.release()
                            except:
                                pass
                        # Skip this camera if it causes issues
                        print(f"[Camera Detection] Error testing index {idx} with {backend_name}: {inner_e}")
                        continue
                            
                except Exception as e:
                    print(f"[Camera Detection] Failed index {idx} with {backend_name}: {e}")
                    continue
            
            # If we found an excellent camera, stop scanning other backends
            if excellent_camera_found:
                break
        
        # Filter and prioritize cameras
        # Note: identify_camera_type already checks OBS running status and marks cameras as 'obs_virtual'
        obs_cameras = []
        usb_cameras = []
        other_cameras = []
        
        for cam in found_cameras:
            if cam['type'] == 'obs_virtual':
                # Camera was identified as OBS Virtual Camera
                obs_cameras.append(cam)
            elif cam['type'] == 'usb_capture':
                # USB capture device (HDMI to USB capture card)
                usb_cameras.append(cam)
            else:
                other_cameras.append(cam)
        
        # Select best camera - always return something if cameras were found
        if obs_cameras:
            # Prefer DSHOW for OBS on Windows (more reliable)
            dshow_obs = [c for c in obs_cameras if c['backend_name'] == 'DSHOW']
            msmt_obs = [c for c in obs_cameras if c['backend_name'] == 'MSMF']
            # Try DSHOW first, then MSMF
            if dshow_obs:
                selected = dshow_obs[0]
            elif msmt_obs:
                selected = msmt_obs[0]
            else:
                selected = obs_cameras[0]
            print(f"[Camera Detection] ✓ Using OBS Virtual Camera (detected): index={selected['index']}, backend={selected['backend_name']}, size={selected['size']}")
            return selected['index'], selected['backend'], selected['backend_name']
        elif usb_cameras:
            # Prefer higher resolution USB capture devices (likely HDMI to USB capture)
            # Also prefer DSHOW backend for USB capture on Windows
            dshow_usb = [c for c in usb_cameras if c['backend_name'] == 'DSHOW']
            if dshow_usb:
                usb_cameras_sorted = sorted(dshow_usb, key=lambda x: (x['width'], x['height']), reverse=True)
                selected = usb_cameras_sorted[0]
            else:
                usb_cameras_sorted = sorted(usb_cameras, key=lambda x: (x['width'], x['height']), reverse=True)
                selected = usb_cameras_sorted[0]
            print(f"[Camera Detection] ✓ Using USB capture device: index={selected['index']}, backend={selected['backend_name']}, size={selected['size']}")
            return selected['index'], selected['backend'], selected['backend_name']
        elif other_cameras:
            # Fall back to other cameras (prefer higher resolution)
            # Prefer DSHOW backend on Windows
            dshow_other = [c for c in other_cameras if c['backend_name'] == 'DSHOW']
            if dshow_other:
                other_sorted = sorted(dshow_other, key=lambda x: (x.get('width', 0), x.get('height', 0)), reverse=True)
                selected = other_sorted[0] if other_sorted else dshow_other[0]
            else:
                other_sorted = sorted(other_cameras, key=lambda x: (x.get('width', 0), x.get('height', 0)), reverse=True)
                selected = other_sorted[0] if other_sorted else other_cameras[0]
            print(f"[Camera Detection] ⚠ Using fallback camera: index={selected['index']}, backend={selected['backend_name']}, size={selected['size']}")
            return selected['index'], selected['backend'], selected['backend_name']
        
        # If no cameras found at all, return default but log warning
        print("[Camera Detection] ⚠ No cameras found in scan, using defaults (0, DSHOW)")
        print("[Camera Detection] ⚠ This usually means no cameras are available or all cameras failed to open")
        return 0, cv2.CAP_DSHOW, "DSHOW"
        
    elif platform.system() == "Darwin":  # macOS
        backend_code = cv2.CAP_AVFOUNDATION
        backend_name = "AVFOUNDATION"
        
        # Check if OBS is running first
        obs_running = False
        try:
            result = subprocess.run(['pgrep', '-f', 'obs'], capture_output=True, timeout=2)
            obs_running = result.returncode == 0
            if obs_running:
                print("[Camera Detection] ✓ OBS is running - prioritizing OBS Virtual Camera indices")
        except:
            pass
        
        # Scan all cameras and identify them
        found_cameras = []
        
        # If OBS is running, prioritize indices 1, 2, 3 (OBS Virtual Camera is usually at 1 or 2)
        # Otherwise, scan normally
        scan_order = [1, 2, 3, 0] + list(range(4, 31)) if obs_running else list(range(31))
        
        for idx in scan_order:
            try:
                print(f"[Camera Detection] Testing index {idx}...")
                cap = cv2.VideoCapture(idx, backend_code)
                if cap.isOpened():
                    # Read multiple frames to verify stability (prevent flickering detection)
                    stable = True
                    first_shape = None
                    for _ in range(3):
                        ret, frame = cap.read()
                        if ret and frame is not None:
                            if first_shape is None:
                                first_shape = frame.shape
                            elif frame.shape != first_shape:
                                stable = False
                                break
                        else:
                            stable = False
                            break
                    
                    if stable and first_shape is not None:
                        h, w = first_shape[:2]
                        camera_type = identify_camera_type(cap, idx)
                        found_cameras.append({
                            'index': idx,
                            'type': camera_type,
                            'size': f"{w}x{h}"
                        })
                        print(f"[Camera Detection] Found camera {idx}: {camera_type} ({w}x{h})")
                    cap.release()
            except Exception as e:
                print(f"[Camera Detection] Failed index {idx}: {e}")
                continue
        
        # Prioritize OBS Virtual Camera
        obs_cameras = [cam for cam in found_cameras if cam['type'] == 'obs_virtual']
        usb_cameras = [cam for cam in found_cameras if cam['type'] == 'usb_capture']
        other_cameras = [cam for cam in found_cameras if cam['type'] in ['facetime', 'other', 'unknown']]
        
        if obs_cameras:
            # Use the first OBS Virtual Camera found (should be index 1 or 2)
            selected = obs_cameras[0]
            print(f"[Camera Detection] ✓ Using OBS Virtual Camera: index={selected['index']}, size={selected['size']}")
            return selected['index'], backend_code, backend_name
        elif usb_cameras:
            # Fall back to USB capture devices
            selected = usb_cameras[0]
            print(f"[Camera Detection] ⚠ Using USB capture: index={selected['index']}, size={selected['size']}")
            return selected['index'], backend_code, backend_name
        elif other_cameras:
            # Fall back to other cameras, but prefer higher indices (less likely to be built-in)
            # Sort by index (higher indices first) to prefer external/OBS cameras
            other_cameras_sorted = sorted(other_cameras, key=lambda x: x['index'], reverse=True)
            selected = other_cameras_sorted[0]
            print(f"[Camera Detection] ⚠ OBS Virtual Camera not found, using {selected['type']}: index={selected['index']}, size={selected['size']}")
            return selected['index'], backend_code, backend_name
        else:
            print("[Camera Detection] ⚠ No cameras found, using defaults (0, AVFOUNDATION)")
            return 0, backend_code, backend_name
        
    else:  # Linux
        backend_code = cv2.CAP_V4L2
        backend_name = "V4L2"
        
        for idx in range(6):
            try:
                cap = cv2.VideoCapture(idx, backend_code)
                if cap.isOpened():
                    ret, frame = cap.read()
                    if ret and frame is not None:
                        cap.release()
                        print(f"[Camera Detection] ✓ Found working camera: index={idx}")
                        return idx, backend_code, backend_name
                    cap.release()
            except Exception:
                continue
        
        print("[Camera Detection] ⚠ No cameras found, using defaults (0, V4L2)")
        return 0, backend_code, backend_name

# New endpoint: enumerate available camera indices/backends
@app.get("/enumerate")
def enumerate_caps():
    results = []
    def try_open(idx, backend, bname):
        cap = cv2.VideoCapture(idx, backend)
        ok = cap.isOpened()
        size = None
        if ok:
            ok2, f = cap.read()
            if ok2 and f is not None:
                h, w = f.shape[:2]
                size = f"{w}x{h}"
            cap.release()
        return {"index": idx, "backend": bname, "opened": bool(ok), "size": size}
    backends = []
    if platform.system()=="Windows":
        backends = [("MSMF", cv2.CAP_MSMF), ("DSHOW", cv2.CAP_DSHOW)]
    elif platform.system()=="Darwin":
        backends = [("AVFOUNDATION", cv2.CAP_AVFOUNDATION)]
    else:
        backends = [("V4L2", cv2.CAP_V4L2)]
    for name, b in backends:
        for i in range(0, 6):
            results.append(try_open(i, b, name))
    return jsonify(results=results)

# New endpoint: auto-detect camera
@app.get("/auto_detect_camera")
def auto_detect_camera_endpoint():
    """Endpoint to trigger camera auto-detection"""
    try:
        idx, backend_code, backend_name = auto_detect_camera()
        # Handle both integer indices and string names (for OBS Virtual Camera)
        index_value = idx if isinstance(idx, (int, str)) else str(idx)
        return jsonify(ok=True, index=index_value, backend=backend_name)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[auto_detect_camera_endpoint] Error: {e}\n{error_detail}")
        return jsonify(ok=False, error=str(e))

# New endpoint: list all available cameras with types
@app.get("/list_cameras")
def list_cameras():
    """List all available cameras with their types"""
    try:
        cameras = []
        
        if platform.system() == "Windows":
            # Windows: Try both DSHOW and MSMF backends
            backends = [
                ("DSHOW", cv2.CAP_DSHOW),
                ("MSMF", cv2.CAP_MSMF)
            ]
        elif platform.system() == "Darwin":
            backends = [("AVFOUNDATION", cv2.CAP_AVFOUNDATION)]
        else:
            backends = [("V4L2", cv2.CAP_V4L2)]
        
        # Limit scan to 0-10 to prevent hangs
        max_scan = 10
        
        for backend_name, backend_code in backends:
            for idx in range(max_scan + 1):
                try:
                    cap = cv2.VideoCapture(idx, backend_code)
                    if cap.isOpened():
                        # Set buffer to 1 for faster response
                        try:
                            cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
                        except:
                            pass
                        
                        ret, frame = cap.read()
                        if ret and frame is not None:
                            h, w = frame.shape[:2]
                            camera_type = identify_camera_type(cap, idx)
                            cameras.append({
                                'index': idx,
                                'backend': backend_name,
                                'type': camera_type,
                                'size': f"{w}x{h}",
                                'width': w,
                                'height': h,
                                'available': True
                            })
                        cap.release()
                except Exception as e:
                    # Skip cameras that cause errors
                    continue
        
        return jsonify(ok=True, cameras=cameras)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[list_cameras] Error: {e}\n{error_detail}")
        return jsonify(ok=False, error=str(e))

# New endpoint: system camera detection
@app.get("/system_cameras")
def system_cameras():
    """Get comprehensive camera information using system commands"""
    try:
        cameras_info = []
        
        if platform.system() == "Windows":
            # Windows: Use PowerShell to get camera information
            try:
                # Get DirectShow devices
                result = subprocess.run([
                    'powershell', '-Command', 
                    'Get-WmiObject -Class Win32_PnPEntity | Where-Object {$_.Name -like "*camera*" -or $_.Name -like "*webcam*" -or $_.Name -like "*video*"} | Select-Object Name, DeviceID'
                ], capture_output=True, text=True, timeout=10)
                
                if result.returncode == 0:
                    cameras_info.append({
                        'method': 'Windows WMI',
                        'output': result.stdout,
                        'error': result.stderr
                    })
            except Exception as e:
                cameras_info.append({
                    'method': 'Windows WMI',
                    'error': str(e)
                })
                
            # Try Device Manager approach
            try:
                result = subprocess.run([
                    'powershell', '-Command',
                    'Get-PnpDevice -Class Camera | Select-Object FriendlyName, InstanceId'
                ], capture_output=True, text=True, timeout=10)
                
                if result.returncode == 0:
                    cameras_info.append({
                        'method': 'Windows PnP',
                        'output': result.stdout,
                        'error': result.stderr
                    })
            except Exception as e:
                cameras_info.append({
                    'method': 'Windows PnP',
                    'error': str(e)
                })
                
        elif platform.system() == "Darwin":  # macOS
            # macOS: Use system_profiler
            try:
                result = subprocess.run([
                    'system_profiler', 'SPCameraDataType'
                ], capture_output=True, text=True, timeout=10)
                
                cameras_info.append({
                    'method': 'macOS System Profiler',
                    'output': result.stdout,
                    'error': result.stderr
                })
            except Exception as e:
                cameras_info.append({
                    'method': 'macOS System Profiler',
                    'error': str(e)
                })
                
        else:  # Linux
            # Linux: Use v4l2-ctl
            try:
                result = subprocess.run([
                    'v4l2-ctl', '--list-devices'
                ], capture_output=True, text=True, timeout=10)
                
                cameras_info.append({
                    'method': 'Linux V4L2',
                    'output': result.stdout,
                    'error': result.stderr
                })
            except Exception as e:
                cameras_info.append({
                    'method': 'Linux V4L2',
                    'error': str(e)
                })
        
        return jsonify(ok=True, system_cameras=cameras_info)
        
    except Exception as e:
        return jsonify(ok=False, error=str(e))

@app.post("/reset_camera_system")
def reset_camera_system_endpoint():
    """Reset the camera system to recover from failed connections."""
    try:
        reset_camera_system()
        return jsonify(ok=True, message="Camera system reset complete")
    except Exception as e:
        return jsonify(ok=False, error=str(e))

@app.post("/force_restart_server")
def force_restart_server():
    """Force restart the Flask server to completely reset everything."""
    try:
        print("[force_restart_server] Initiating server restart...")
        
        # Stop all video threads and reset camera system
        reset_camera_system()
        
        # Create a restart script and execute it
        import os
        import sys
        import subprocess
        
        if platform.system() == "Windows":
            # Windows: Create a batch file to restart
            restart_script = "restart_app.bat"
            with open(restart_script, 'w') as f:
                f.write(f"@echo off\n")
                f.write(f"timeout /t 2 /nobreak >nul\n")
                f.write(f'"{sys.executable}" {" ".join(sys.argv)}\n')
            
            # Execute the batch file
            subprocess.Popen([restart_script], shell=True)
        else:
            # Unix/Linux/macOS: Use shell script
            restart_script = "restart_app.sh"
            with open(restart_script, 'w') as f:
                f.write("#!/bin/bash\n")
                f.write("sleep 2\n")
                f.write(f'"{sys.executable}" {" ".join(sys.argv)}\n')
            
            os.chmod(restart_script, 0o755)
            subprocess.Popen([f"./{restart_script}"], shell=True)
        
        # Exit current process
        os._exit(0)
        
    except Exception as e:
        print(f"[force_restart_server] Error: {e}")
        return jsonify(ok=False, error=str(e))

# New endpoint: probe source and show info for selected backend/fourcc/res
@app.post("/probe")
def probe():
    cfg = request.get_json(force=True) if request.is_json else {}
    src = cfg.get("src", mon.src)
    backend_name = cfg.get("backend", mon.backend_name)
    fourcc_sel = (cfg.get("fourcc", mon.fourcc) or "auto").upper()
    res_preset = (cfg.get("res", mon.res_preset) or "1080p").lower()
    stream = cfg.get("stream", mon.stream_path) or ""

    # Determine desired resolution
    if res_preset == "720p":
        want_w, want_h = 1280, 720
    else:
        want_w, want_h = 1920, 1080

    # Helper to open and read one frame with options
    def try_one(_src, _backend, _bname):
        cap = None
        opened = False
        size = None
        fourcc_read = ''
        fps_val = None
        try:
            if isinstance(_src, str) and not _src.isdigit():
                cap = cv2.VideoCapture(_src)
            else:
                try:
                    idx = int(_src)
                except Exception:
                    idx = _src
                cap = cv2.VideoCapture(idx, _backend)
            if not cap or not cap.isOpened():
                return {"index": _src, "backend": _bname, "opened": False}

            # Normalize & apply user preferences
            try: cap.set(cv2.CAP_PROP_CONVERT_RGB, 1)
            except Exception: pass
            if fourcc_sel == "MJPG":
                try: cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'MJPG'))
                except Exception: pass
            elif fourcc_sel == "YUY2":
                try: cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'YUY2'))
                except Exception: pass
            try:
                cap.set(cv2.CAP_PROP_FRAME_WIDTH,  want_w)
                cap.set(cv2.CAP_PROP_FRAME_HEIGHT, want_h)
            except Exception:
                pass
            try: cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
            except Exception: pass
            try: cap.set(cv2.CAP_PROP_FPS, 30)
            except Exception: pass

            opened = cap.isOpened()
            if not opened:
                return {"index": _src, "backend": _bname, "opened": False}

            # Probe a few frames
            for _ in range(5):
                cap.read()
            ok, frame = cap.read()
            if ok and frame is not None:
                h, w = frame.shape[:2]
                size = f"{w}x{h}"
            # Read back FOURCC/FPS if possible
            try:
                v = cap.get(cv2.CAP_PROP_FOURCC)
                if v and v != 0:
                    fourcc_read = _fourcc_to_str(int(v))
            except Exception:
                pass
            try:
                fps_g = cap.get(cv2.CAP_PROP_FPS)
                if fps_g and fps_g > 0:
                    fps_val = round(float(fps_g), 1)
            except Exception:
                pass
            return {"index": _src, "backend": _bname, "opened": True, "size": size, "fourcc": fourcc_read, "fps": fps_val}
        finally:
            try:
                if cap: cap.release()
            except Exception:
                pass

    # If stream/URL provided, probe just that without backend variations
    if isinstance(stream, str) and stream.strip():
        res = [try_one(stream.strip(), 0, "AUTO(URL)")]
        return jsonify(results=res, note="Probed stream URL")

    # Otherwise probe common backends per-OS
    backends = []
    if platform.system() == "Windows":
        # If user chose a specific backend, test it first then the alternate
        name = (backend_name or "auto").lower()
        if name == "msmf":
            backends = [("MSMF", cv2.CAP_MSMF), ("DSHOW", cv2.CAP_DSHOW)]
        elif name == "dshow":
            backends = [("DSHOW", cv2.CAP_DSHOW), ("MSMF", cv2.CAP_MSMF)]
        else:
            backends = [("MSMF", cv2.CAP_MSMF), ("DSHOW", cv2.CAP_DSHOW)]
    elif platform.system() == "Darwin":
        backends = [("AVFOUNDATION", cv2.CAP_AVFOUNDATION)]
    else:
        backends = [("V4L2", cv2.CAP_V4L2)]

    results = []
    for bname, bconst in backends:
        results.append(try_one(src, bconst, bname))
    return jsonify(results=results)


# Helper function to begin monitoring with a config dict
def _begin_with_cfg(cfg: dict):
    with mon.lock:
        if mon.running:
            # already running: report current CSV so UI can show it
            return jsonify(ok=True, already=True, csv=mon.csv_path)

        mon.last_cfg = dict(cfg) if isinstance(cfg, dict) else {}
        mon.src          = cfg.get("src", mon.src)
        mon.backend_name = cfg.get("backend", mon.backend_name)
        mon.fourcc       = cfg.get("fourcc", mon.fourcc)
        mon.res_preset   = cfg.get("res", mon.res_preset)
        mon.stream_path  = cfg.get("stream", mon.stream_path)
        mon.center_w     = int(cfg.get("cw", mon.center_w))
        mon.thresh       = int(cfg.get("thr", mon.thresh))
        mon.stable_frames= int(cfg.get("st", mon.stable_frames))
        mon.dark_mean    = float(cfg.get("dm", mon.dark_mean))
        mon.dark_std     = float(cfg.get("ds", mon.dark_std))
        # Optional overrides from UI
        mon.white_frac_gate = cfg.get("white_pct", mon.white_frac_gate)
        mon.mean_gate       = cfg.get("mean_gate", mon.mean_gate)
        try:
            rin = cfg.get("roi_inset", None)
            if rin is not None:
                mon.roi_inset_px = int(rin or 0)
        except Exception:
            pass
        # Prefer provided paths only if they exist; otherwise fall back to bundled art/
        req_bars = cfg.get("bars", "")
        req_intr = cfg.get("intr", "")
        req_bars = req_bars.strip() if isinstance(req_bars, str) else ""
        req_intr = req_intr.strip() if isinstance(req_intr, str) else ""
        mon.bars_ref = req_bars if (req_bars and os.path.exists(req_bars)) else BARS_REF
        mon.int_ref  = req_intr if (req_intr and os.path.exists(req_intr)) else INT_REF
        # Build list of usable INT reference images (requested first)
        refs = []
        if mon.int_ref and os.path.exists(mon.int_ref):
            refs.append(mon.int_ref)
        for p in discover_int_refs():
            if p not in refs:
                refs.append(p)
        mon.int_ref_paths = refs

        # Always roll a fresh CSV on every start (ignore any incoming csv path)
        mon.reset_counts_and_roll_csv()
        new_csv = mon.csv_path

        mon.running = True
        mon.status = "starting"
        mon.last_frame = None
        mon.last_crop  = None
        mon.last_metrics = {"db": None, "di": None, "mean": None, "std": None}
        mon.await_connect = True

    t = threading.Thread(target=run_loop, daemon=True)
    mon.worker = t
    t.start()
    # Tell the UI which CSV is active
    return jsonify(ok=True, csv=new_csv, rolled=True)

# POST /start: Accepts JSON, starts monitoring with config
@app.post("/start")
def start():
    # Accept empty or malformed JSON gracefully
    cfg = request.get_json(silent=True) or {}
    print("POST /start called with cfg:", cfg)
    return _begin_with_cfg(cfg)

# GET /start: Alias for starting with current settings (no body)
@app.get("/start")
def start_get():
    print("GET /start called (no body); starting with current settings")
    return _begin_with_cfg({})

@app.post("/update_camera_settings")
def update_camera_settings():
    """Update camera settings from UI."""
    cfg = request.get_json(silent=True) or {}
    
    with mon.lock:
        if "src" in cfg:
            try:
                # Try integer first, fall back to string (for OBS Virtual Camera names)
                mon.src = int(cfg["src"])
            except:
                mon.src = cfg["src"]  # Could be string path or camera name
        
        if "backend" in cfg:
            backend_name = cfg["backend"].upper()
            mon.backend_name = backend_name
        
        if "fourcc" in cfg:
            mon.fourcc = cfg["fourcc"]
        
        if "res" in cfg:
            mon.res_preset = cfg["res"]
    
    print(f"[update_camera_settings] Updated: Source={mon.src}, Backend={mon.backend_name}, FOURCC={mon.fourcc}, Res={mon.res_preset}")
    return jsonify(ok=True)

@app.post("/start_video")
def start_video():
    """Start or restart video feed with current camera settings."""
    try:
        # Stop existing video if running
        with mon.lock:
            if mon.video_running:
                print("[start_video] Stopping existing video feed...")
                mon.video_running = False
                mon.detection_active = False
                mon.running = False
                # Wait a bit for the thread to stop
                import time
                time.sleep(0.5)
            
            mon.last_cfg = {}
            mon.video_running = True
            mon.detection_active = False
            mon.running = True  # for compatibility
            mon.status = "video_starting"
            mon.last_frame = None
            mon.last_crop = None

        print(f"[start_video] Starting video with Source={mon.src}, Backend={mon.backend_name}")
        t = threading.Thread(target=run_loop, daemon=True)
        mon.worker = t
        t.start()
        return jsonify(ok=True)
        
    except Exception as e:
        print(f"[start_video] ERROR: {e}")
        with mon.lock:
            mon.video_running = False
            mon.detection_active = False
            mon.running = False
            mon.status = f"error: {e}"
        return jsonify(ok=False, error=str(e))

@app.post("/start_detection")
def start_detection():
    """Engage detection and logging (video must already be running)."""
    # Wait for video to be ready (with timeout)
    import time
    max_wait = 10.0  # Maximum seconds to wait for video
    wait_interval = 0.1  # Check every 100ms
    waited = 0.0
    
    while waited < max_wait:
        with mon.lock:
            if mon.video_running:
                break
            status = mon.status
        time.sleep(wait_interval)
        waited += wait_interval
    
    # Check again after waiting
    with mon.lock:
        if not mon.video_running:
            # Try to start video if not running
            print("[Start Test] Video not running, attempting to start video feed...")
            try:
                mon.video_running = True
                mon.running = True
                mon.status = "video_starting"
                mon.last_frame = None
                mon.last_crop = None
                
                # Start video thread
                t = threading.Thread(target=run_loop, daemon=True)
                mon.worker = t
                t.start()
                print("[Start Test] Video feed thread started")
                
                # Wait a bit for video to initialize
                time.sleep(0.5)
                
                # Wait a bit more and check again
                time.sleep(1.0)
                if not mon.video_running or mon.status in ["error", "error: cannot read from source", "error: cannot open video source"]:
                    error_msg = f"Video failed to start. Status: {mon.status}. Please check camera connection and try 'Connect to Camera' button."
                    print(f"[Start Test] {error_msg}")
                    return jsonify(ok=False, error=error_msg), 400
            except Exception as e:
                print(f"[Start Test] Error starting video: {e}")
                import traceback
                traceback.print_exc()
                return jsonify(ok=False, error=f"Failed to start video: {e}. Please try 'Connect to Camera' button first."), 500
        
        if mon.detection_active:
            return jsonify(ok=True, already=True)
        
        # Reset counters and start fresh CSV with headers
        mon.reset_counts_and_roll_csv()
        
        # Set detection_active BEFORE logging initial states so _csv_append will work
        mon.detection_active = True
        mon.status = "detection_starting"
        csv_path = mon.csv_path
        print(f"[Start Test] Set detection_active=True, status={mon.status}")
    
    # Do I/O operations outside the lock to prevent blocking
    try:
        # Create CSV file with headers immediately when Start Test is pressed
        f, w = open_csv(csv_path)
        f.close()
        print(f"[Start Test] Created CSV with headers: {csv_path}")
    except Exception as e:
        print(f"[Start Test] Error creating CSV: {e}")
        with mon.lock:
            mon.detection_active = False
        return jsonify(ok=False, error=f"Failed to create CSV: {e}"), 500
    
    # Log initial "No Signal" state for all 6 channels (outside lock)
    # This ensures the first entry is always "No Signal" for proper elapsed time calculations
    TILE_TO_VIDEO = [1, 3, 5, 2, 4, 6]  # Map tile indices to video numbers
    try:
        for video_num in TILE_TO_VIDEO:
            _csv_append(
                video=video_num,
                state="NO_SIGNAL",
                elapsed_secs=None,  # No elapsed time for initial state (and NO_SIGNAL never has elapsed_secs)
                cycle_num=None,
                partial_cycle_num=None,
                bars_dist=None,
                int_dist=None,
                event_type="test_start"
            )
        print(f"[Start Test] Logged initial 'No Signal' state for all 6 channels")
    except Exception as e:
        print(f"[Start Test] Warning: Error logging initial No Signal states: {e}")
    
    # Verify detection_active is still True before returning
    with mon.lock:
        print(f"[Start Test] Returning response: detection_active={mon.detection_active}, csv_path={csv_path}")
        detection_active = mon.detection_active
    
    return jsonify(ok=True, csv_path=csv_path, detection_active=detection_active)

@app.post("/end_test")
def end_test():
    """End the test, recalculate elapsed times, generate report, and append to CSV."""
    with mon.lock:
        mon.detection_active = False
        mon.status = "generating_report"
    
    # Post-process: Recalculate elapsed times from previous No Signal for each channel
    try:
        print("[End Test] Recalculating elapsed times from No Signal...")
        _recalculate_elapsed_times_from_no_signal()
    except Exception as e:
        print(f"[End Test] Warning: Error recalculating elapsed times: {e}")
    
    # Generate report and append to CSV
    try:
        report_text, csv_path = _generate_and_append_report()
        print("\n" + report_text)  # Print to console
        print(f"[End Test] Report generated successfully, {len(report_text)} chars")
        
        return jsonify(ok=True, report=report_text, csv_path=csv_path)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[End Test] Error generating report: {e}")
        print(error_detail)
        return jsonify(ok=False, error=str(e), detail=error_detail)

@app.post("/reveal_csv")
def reveal_csv():
    """Open the file browser and highlight the most recent CSV and Excel files."""
    try:
        # Get the current CSV path
        with mon.lock:
            csv_path = mon.csv_path
        
        # Reveal CSV in file browser
        success_csv = _reveal_in_file_browser(csv_path)
        
        # Also reveal Excel file if it exists
        excel_path = _get_excel_path(csv_path)
        success_excel = False
        if EXCEL_AVAILABLE and os.path.exists(excel_path):
            success_excel = _reveal_in_file_browser(excel_path)
        
        return jsonify(ok=True, success=success_csv, csv_path=csv_path, excel_path=excel_path if success_excel else None)
    except Exception as e:
        print(f"[Reveal CSV] Error: {e}")
        return jsonify(ok=False, error=str(e))

@app.post("/clear")
def clear():
    with mon.lock:
        if mon.running:
            return jsonify(error="stop first"), 400
        mon.reset_counts_and_roll_csv()
    return jsonify(ok=True, csv=mon.csv_path)

# New endpoint: reset tallies (counters and last-seen) without rolling CSV, allowed even while running
@app.post("/reset_tallies")
def reset_tallies():
    with mon.lock:
        mon.reset_tallies()
    return jsonify(ok=True)

@app.get("/download")
def download_csv():
    path = os.path.abspath(mon.csv_path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    if not os.path.isfile(path):
        return jsonify(error="CSV not found", path=path), 404
    return send_file(path, as_attachment=True, download_name=os.path.basename(path))


# Optional: favicon route to suppress 404 noise
@app.get("/favicon.ico")
def favicon():
    return Response(status=204)

@app.post("/shutdown")
def shutdown():
    # Graceful quit from the web UI
    def _exit():
        time.sleep(0.3)
        os._exit(0)
    threading.Thread(target=_exit, daemon=True).start()
    return jsonify(ok=True)



def open_browser():
    global PORT
    url = f"http://localhost:{PORT}/"
    try:
        if platform.system()=="Windows":
            os.startfile(url)  # type: ignore[attr-defined]
        else:
            webbrowser.open(url)
    except Exception:
        pass
    print(f"Open {url} in your browser if it didn't open automatically.")

# ---- main launcher ----
if __name__ == "__main__":
    # Auto-detect and connect to camera on startup (non-blocking with timeout)
    def auto_connect_camera():
        """Auto-detect and connect to camera in background thread."""
        try:
            import time
            # Small delay to ensure server is fully started
            time.sleep(0.5)
            print("[Startup] Auto-detecting camera...")
            
            # On Windows, check if OBS is running first to prioritize OBS Virtual Camera
            if platform.system() == "Windows":
                obs_running = False
                try:
                    result = subprocess.run(['tasklist', '/FI', 'IMAGENAME eq obs64.exe'], 
                                          capture_output=True, timeout=2)
                    obs_running = result.returncode == 0 and b'obs64.exe' in result.stdout
                    if obs_running:
                        print("[Startup] ✓ OBS is running - will prioritize OBS Virtual Camera")
                    else:
                        print("[Startup] ⚠ OBS is not running - will use any available camera")
                except Exception as e:
                    print(f"[Startup] Could not check OBS status: {e}")
            
            detected_idx, detected_backend_code, detected_backend_name = auto_detect_camera()
            
            with mon.lock:
                mon.src = detected_idx
                mon.backend_name = detected_backend_name
            print(f"[Startup] ✓ Camera detected: Source={detected_idx}, Backend={detected_backend_name}")
            
            # Automatically start video feed with detected camera
            print("[Startup] Auto-connecting to camera...")
            with mon.lock:
                if not mon.video_running:  # Only start if not already running
                    mon.video_running = True
                    mon.running = True
                    mon.status = "video_starting"
                    mon.last_frame = None
                    mon.last_crop = None
                    
                    # Start video thread
                    t = threading.Thread(target=run_loop, daemon=True)
                    mon.worker = t
                    t.start()
                    print(f"[Startup] ✓ Video feed thread started with Source={detected_idx}, Backend={detected_backend_name}")
                    
                    # Give video thread a moment to initialize
                    time.sleep(1.0)
                    
                    # Verify video is actually running
                    with mon.lock:
                        if mon.video_running and mon.status not in ["error", "error: cannot read from source", "error: cannot open video source"]:
                            print(f"[Startup] ✓ Video feed confirmed running: status={mon.status}")
                        else:
                            print(f"[Startup] ⚠ Video feed may have failed: status={mon.status}, video_running={mon.video_running}")
                else:
                    print(f"[Startup] ⚠ Video feed already running, skipping auto-connect")
        except Exception as e:
            print(f"[Startup] ⚠ Auto-detection/connection failed: {e}, using defaults")
            import traceback
            traceback.print_exc()
            # Set defaults if auto-connect fails
            with mon.lock:
                if not mon.video_running:
                    print("[Startup] Attempting to start with default camera settings...")
                    try:
                        mon.video_running = True
                        mon.running = True
                        mon.status = "video_starting"
                        mon.last_frame = None
                        mon.last_crop = None
                        t = threading.Thread(target=run_loop, daemon=True)
                        mon.worker = t
                        t.start()
                        print("[Startup] Default video feed thread started")
                    except Exception as e2:
                        print(f"[Startup] Failed to start default video feed: {e2}")
    
    # Start auto-connection in background thread so server can start immediately
    # This prevents the server from hanging if camera detection is slow
    auto_connect_thread = threading.Thread(target=auto_connect_camera, daemon=True)
    auto_connect_thread.start()
    
    # Pick a port (try 5055 first, otherwise the next available up to 5070)
    PORT = _choose_port(5055, 5070)
    print(f"Starting server on http://localhost:{PORT} …")
    # Launch the browser shortly after the server starts
    threading.Timer(0.6, open_browser).start()
    app.run(host="127.0.0.1", port=PORT, debug=False)